"""
Kira Web Runtime API

Streamlit UI 없이 웹 랜딩(index.html)에서 Kira 분석 엔진을 직접 실행하기 위한 API 서버.
"""

# Do NOT add 'from __future__ import annotations' — it breaks FastAPI's
# runtime introspection of list[UploadFile] parameters (ForwardRef error).

import asyncio
import base64
import hashlib
import hmac as hmac_mod
import json
import logging
import os
import re
import secrets
import shutil
import sys
import tempfile
import traceback
import urllib.error
import urllib.parse
import urllib.request
import uuid
import threading
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta, timezone
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Any

import httpx

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded


ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
_WEB_APP_DIR = str(Path(__file__).resolve().parent)
if _WEB_APP_DIR not in sys.path:
    sys.path.insert(0, _WEB_APP_DIR)

load_dotenv(ROOT_DIR / ".env")

# Only init Bid Workspace DB if BID_DATABASE_URL is set (gradual rollout)
_BID_DB_ENABLED = bool(os.getenv("BID_DATABASE_URL"))

logger = logging.getLogger(__name__)

from engine import RAGEngine  # noqa: E402
from nara_api import (  # noqa: E402
    search_bids as nara_search_bids,
    search_order_plans as nara_search_order_plans,
    get_bid_detail_attachments as nara_get_bid_detail_attachments,
    get_bid_attachments as nara_get_bid_attachments,
    download_attachment as nara_download_attachment,
    pick_best_attachment,
    CATEGORY_CODE as NARA_CATEGORY_CODE,
)
from matcher import MatchStatus, MatchingResult, QualificationMatcher  # noqa: E402
from rfx_analyzer import RFxAnalysisResult, RFxAnalyzer  # noqa: E402
from document_parser import DocumentParser  # noqa: E402
from chat_tools import CHAT_TOOLS, TOOL_USE_SYSTEM_PROMPT, parse_tool_call_result  # noqa: E402
from services.web_app.react_chat import react_chat_loop  # noqa: E402
from chat_router import (  # noqa: E402
    ChatPolicy,
    RouteDecision,
    RouteIntent,
    UNSAFE_KEYWORDS,
    default_router_log_path,
    write_router_telemetry,
)
from user_store import (  # noqa: E402
    create_user_session,
    enforce_usage_quota,
    get_actor_usage_counts,
    get_usage_overview,
    list_usage_by_actor,
    get_user_profile,
    init_user_store,
    invalidate_user_session,
    resolve_user_from_session,
    upsert_social_user,
    username_to_scope,
)
from alert_storage import get_alert_config, save_alert_config  # noqa: E402
from services.web_app.session_store import (  # noqa: E402
    load_analysis_from_redis,
    save_analysis_to_redis,
    restore_analysis_from_dict,
)


ALLOWED_UPLOAD_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".txt",
    ".hwp", ".hwpx",
    ".xlsx", ".xls", ".csv",
    ".pptx", ".ppt",
}
DEFAULT_SESSION_TTL_HOURS = 12
FRONTEND_DIR = ROOT_DIR / "frontend" / "kirabot"
FRONTEND_DIST_DIR = FRONTEND_DIR / "dist"
GAP_QUERY_KEYWORDS: tuple[str, ...] = (
    "부족",
    "미충족",
    "모자",
    "불충분",
    "결격",
    "참여하기엔",
    "안되는 이유",
    "왜 안",
    "리스크",
    "보완",
)

DEFAULT_SUGGESTIONS_NO_CONTEXT: list[str] = [
    "회사 소개서 PDF를 먼저 올리면 어떤 항목을 확인할 수 있어?",
    "분석할 문서를 업로드하면 어떤 결과가 나오는지 알려줘",
    "이 도구로 우리 회사 입찰 준비 시간을 얼마나 줄일 수 있어?",
    "첫 분석 전에 준비할 파일 2가지를 알려줘",
]

DEFAULT_SUGGESTIONS_COMPANY_ONLY: list[str] = [
    "우리 회사 강점 3가지를 문서 근거로 요약해줘",
    "우리 회사 보유 인증/실적 항목을 정리해줘",
    "입찰 준비에 바로 쓸 수 있는 회사 소개 포인트를 뽑아줘",
    "분석할 공고 문서를 올리기 전에 부족한 자료를 점검해줘",
]

DEFAULT_SUGGESTIONS_WITH_ANALYSIS: list[str] = [
    "이 공고 필수요건 중 미충족 항목만 우선순위로 정리해줘",
    "마감일까지 준비해야 할 서류 체크리스트를 만들어줘",
    "근거 페이지와 함께 GO/NO-GO 판단을 짧게 알려줘",
    "점수에 가장 큰 영향을 주는 항목 3개를 알려줘",
]

OAUTH_STATE_TTL_SECONDS = 600
DEFAULT_CHAT_DAILY_LIMIT = 20
DEFAULT_ANALYZE_MONTHLY_LIMIT = 30


class SessionPayload(BaseModel):
    session_id: str


class CompanyDeletePayload(BaseModel):
    session_id: str
    source_file: str


class AnalyzeTextPayload(BaseModel):
    session_id: str
    document_text: str


class ChatPayload(BaseModel):
    session_id: str
    message: str
    source_files: list[str] | None = None


class BidSearchPayload(BaseModel):
    keywords: list[str] | str | None = None
    category: str = "all"
    region: str = ""
    minAmt: float | None = None
    maxAmt: float | None = None
    period: str = "1m"
    excludeExpired: bool = True
    page: int = Field(default=1, ge=1, le=1000)
    pageSize: int = Field(default=20, ge=1, le=100)


class BidAnalyzePayload(BaseModel):
    session_id: str
    bid_ntce_no: str
    bid_ntce_ord: str = "00"
    category: str = ""


class BidEvaluateBatchPayload(BaseModel):
    session_id: str
    bid_ntce_nos: list[str] = Field(max_length=50)


class SmartFitScorePayload(BaseModel):
    session_id: str
    bid_notice_id: str = ""
    bid_title: str = ""
    keywords: list[str] = []


@dataclass
class WebRuntimeSession:
    """웹 UI 세션별 RAG 엔진 컨텍스트."""

    session_id: str
    rag_engine: RAGEngine
    rfx_rag_engine: RAGEngine
    created_at: datetime
    last_used_at: datetime
    latest_rfx_analysis: RFxAnalysisResult | None = None
    latest_matching_result: MatchingResult | None = None
    latest_document_name: str = ""
    _inject_lock: threading.Lock = field(default_factory=threading.Lock, repr=False)


@asynccontextmanager
async def lifespan(app: FastAPI):
    if _BID_DB_ENABLED:
        from services.web_app.db import init_db, close_db
        await init_db()
        logger.info("Bid Workspace DB initialized")

    task = asyncio.create_task(_alert_scheduler_loop())
    logger.info("Alert scheduler started")
    yield
    task.cancel()
    try:
        await task
    except asyncio.CancelledError:
        pass

    if _BID_DB_ENABLED:
        from services.web_app.db import close_db
        await close_db()
        logger.info("Bid Workspace DB closed")


# Rate limiting — global default + Studio-specific
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["60/minute"],  # Global default: 60 req/min per IP
)

app = FastAPI(title="Kira Web Runtime", version="0.1.0", lifespan=lifespan)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# SlowAPIMiddleware enforces default_limits on ALL routes (including Studio)
from slowapi.middleware import SlowAPIMiddleware
app.add_middleware(SlowAPIMiddleware)

# Bid Workspace routers (only when DB is configured)
if _BID_DB_ENABLED:
    from services.web_app.api.projects import router as projects_router
    from services.web_app.api.assets import router as assets_router
    from services.web_app.api.generate import router as generate_router
    from services.web_app.api.studio import router as studio_router
    app.include_router(projects_router)
    app.include_router(assets_router)
    app.include_router(generate_router)
    app.include_router(studio_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        origin.strip()
        for origin in os.getenv(
            "WEB_API_ALLOW_ORIGINS",
            "http://localhost:3000,http://127.0.0.1:3000,http://localhost:5173,http://127.0.0.1:5173,http://localhost:8080,http://127.0.0.1:8080",
        ).split(",")
        if origin.strip()
    ],
    allow_origin_regex=r"^https?://(localhost|127\.0\.0\.1)(:\d+)?$|^https://m-s-solutions-production\.up\.railway\.app$",
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "X-Requested-With"],
)


@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    """Add security headers to all responses."""
    response = await call_next(request)
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains; preload"
    # CSP: allow self + inline styles (Tailwind) + Railway domain + data: for images
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: https:; "
        "connect-src 'self' https://*.up.railway.app https://*.go.kr; "
        "frame-src https://my.spline.design; "
        "frame-ancestors 'none';"
    )
    return response


@app.middleware("http")
async def add_request_id(request: Request, call_next):
    """Add request_id for structured logging correlation."""
    import uuid
    request_id = str(uuid.uuid4())
    request.state.request_id = request_id
    response = await call_next(request)
    response.headers["X-Request-ID"] = request_id
    return response


@app.middleware("http")
async def log_security_events(request: Request, call_next):
    """Log security events for 401/403/409 responses."""
    response = await call_next(request)
    if response.status_code in (401, 403, 409, 429):
        from services.web_app.structured_logger import get_structured_logger
        slog = get_structured_logger("web_app.security")
        slog.set_request_id(getattr(request.state, "request_id", None))
        slog.security(
            "security_event",
            status_code=response.status_code,
            method=request.method,
            path=request.url.path,
            client_ip=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent"),
        )
    return response

SESSIONS: dict[str, WebRuntimeSession] = {}
# HMAC key for signing OAuth state tokens (survives restarts via env var)
_OAUTH_STATE_KEY = (os.getenv("OAUTH_STATE_SECRET") or os.getenv("GOOGLE_OAUTH_CLIENT_SECRET") or "").encode()
if not _OAUTH_STATE_KEY:
    import secrets as _secrets
    _OAUTH_STATE_KEY = _secrets.token_bytes(32)
    logger.warning("OAUTH_STATE_SECRET not set — using ephemeral random key (will break on restart)")

init_user_store()

if (FRONTEND_DIST_DIR / "assets").exists():
    app.mount("/assets", StaticFiles(directory=FRONTEND_DIST_DIR / "assets"), name="frontend-assets")


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _session_ttl() -> timedelta:
    try:
        ttl_hours = int(os.getenv("WEB_SESSION_TTL_HOURS", str(DEFAULT_SESSION_TTL_HOURS)).strip())
    except ValueError:
        ttl_hours = DEFAULT_SESSION_TTL_HOURS
    return timedelta(hours=max(1, ttl_hours))


def _safe_float_env(name: str, default: float) -> float:
    try:
        return float(os.getenv(name, str(default)).strip())
    except ValueError:
        return float(default)


def _safe_bool_env(name: str, default: bool) -> bool:
    raw = os.getenv(name, "1" if default else "0").strip().lower()
    if not raw:
        return default
    return raw not in {"0", "false", "off", "no"}


def _safe_int_env(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)).strip())
    except ValueError:
        return int(default)


def _quota_enabled() -> bool:
    return _safe_bool_env("QUOTA_ENABLED", True)


def _auth_cookie_name() -> str:
    return os.getenv("AUTH_COOKIE_NAME", "kira_auth").strip() or "kira_auth"


def _auth_cookie_path() -> str:
    path = os.getenv("AUTH_COOKIE_PATH", "/").strip()
    return path if path.startswith("/") else "/"


def _auth_cookie_domain() -> str | None:
    value = os.getenv("AUTH_COOKIE_DOMAIN", "").strip()
    return value or None


def _auth_cookie_secure() -> bool:
    return _safe_bool_env("AUTH_COOKIE_SECURE", False)


def _resolve_usage_actor(request: Request, session_id: str) -> tuple[str, str]:
    token = str(request.cookies.get(_auth_cookie_name(), "") or "")
    username = str(resolve_user_from_session(token) or "")
    if username:
        return f"user:{username_to_scope(username)}", username
    return f"anon:{_sanitize_session_id(session_id)}", ""


def _usage_limits_for_action(action: str) -> tuple[int, int]:
    action_norm = str(action or "").strip().lower()
    if action_norm == "chat":
        return (
            max(0, _safe_int_env("QUOTA_CHAT_DAILY_LIMIT", DEFAULT_CHAT_DAILY_LIMIT)),
            max(0, _safe_int_env("QUOTA_CHAT_MONTHLY_LIMIT", 0)),
        )
    if action_norm == "analyze":
        return (
            max(0, _safe_int_env("QUOTA_ANALYZE_DAILY_LIMIT", 0)),
            max(0, _safe_int_env("QUOTA_ANALYZE_MONTHLY_LIMIT", DEFAULT_ANALYZE_MONTHLY_LIMIT)),
        )
    return (0, 0)


def _enforce_quota_or_raise(
    *,
    request: Request,
    session_id: str,
    action: str,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if not _quota_enabled():
        return {
            "enabled": False,
            "action": action,
            "today_used": 0,
            "today_limit": 0,
            "month_used": 0,
            "month_limit": 0,
        }

    actor_key, username = _resolve_usage_actor(request, session_id)
    daily_limit, monthly_limit = _usage_limits_for_action(action)
    decision = enforce_usage_quota(
        actor_key=actor_key,
        username=username,
        action=action,
        daily_limit=daily_limit,
        monthly_limit=monthly_limit,
        metadata=metadata or {},
    )
    decision["enabled"] = True
    if not decision.get("allowed", False):
        detail = {
            "message": "요청 한도를 초과했습니다.",
            "action": action,
            "reason": decision.get("reason", ""),
            "today_used": decision.get("today_used", 0),
            "today_limit": decision.get("today_limit", 0),
            "month_used": decision.get("month_used", 0),
            "month_limit": decision.get("month_limit", 0),
        }
        raise HTTPException(status_code=429, detail=detail)
    return decision


def _admin_usernames() -> set[str]:
    values: set[str] = set()
    for env_name in ("SUPER_ADMIN_USERNAMES", "OPERATOR_USERNAMES", "ADMIN_USERNAMES"):
        raw = os.getenv(env_name, "")
        for item in raw.split(","):
            normalized = item.strip()
            if normalized:
                values.add(normalized)
    return values


def _google_client_id() -> str:
    value = os.getenv("GOOGLE_OAUTH_CLIENT_ID", "").strip()
    if not value:
        raise HTTPException(status_code=503, detail="GOOGLE_OAUTH_CLIENT_ID가 설정되지 않았습니다.")
    return value


def _google_client_secret() -> str:
    value = os.getenv("GOOGLE_OAUTH_CLIENT_SECRET", "").strip()
    if not value:
        raise HTTPException(status_code=503, detail="GOOGLE_OAUTH_CLIENT_SECRET이 설정되지 않았습니다.")
    return value


def _google_redirect_uri() -> str:
    value = os.getenv("GOOGLE_OAUTH_REDIRECT_URI", "").strip()
    if not value:
        raise HTTPException(status_code=503, detail="GOOGLE_OAUTH_REDIRECT_URI가 설정되지 않았습니다.")
    return value


def _google_auth_url() -> str:
    return os.getenv("GOOGLE_OAUTH_AUTH_URL", "https://accounts.google.com/o/oauth2/v2/auth").strip()


def _google_token_url() -> str:
    return os.getenv("GOOGLE_OAUTH_TOKEN_URL", "https://oauth2.googleapis.com/token").strip()


def _google_userinfo_url() -> str:
    return os.getenv("GOOGLE_OAUTH_USERINFO_URL", "https://openidconnect.googleapis.com/v1/userinfo").strip()


def _google_scopes() -> str:
    return os.getenv("GOOGLE_OAUTH_SCOPES", "openid email profile").strip() or "openid email profile"


_ALLOWED_HOSTS: set[str] = set(
    h.strip() for h in os.getenv("ALLOWED_HOSTS", "kirabot.co.kr,www.kirabot.co.kr").split(",") if h.strip()
)


def _post_login_url(request: Request | None = None) -> str:
    """OAuth 완료 후 리다이렉트 URL. 요청의 origin을 자동 감지."""
    explicit = os.getenv("GOOGLE_OAUTH_POST_LOGIN_URL", "").strip()
    if explicit:
        return explicit
    if request:
        host = request.headers.get("host", "").split(":")[0]
        scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
        if host and host in _ALLOWED_HOSTS:
            return f"{scheme}://{host}/"
    return "http://localhost:8000/"


# ── Kakao OAuth Helpers ──

def _kakao_client_id() -> str:
    value = os.getenv("KAKAO_CLIENT_ID", "").strip()
    if not value:
        raise HTTPException(status_code=503, detail="KAKAO_CLIENT_ID가 설정되지 않았습니다.")
    return value


def _kakao_client_secret() -> str:
    return os.getenv("KAKAO_CLIENT_SECRET", "").strip()


def _kakao_redirect_uri() -> str:
    value = os.getenv("KAKAO_REDIRECT_URI", "").strip()
    if not value:
        raise HTTPException(status_code=503, detail="KAKAO_REDIRECT_URI가 설정되지 않았습니다.")
    return value


def _exchange_kakao_code_for_token(code: str) -> dict[str, Any]:
    params: dict[str, str] = {
        "grant_type": "authorization_code",
        "client_id": _kakao_client_id(),
        "redirect_uri": _kakao_redirect_uri(),
        "code": code,
    }
    secret = _kakao_client_secret()
    if secret:
        params["client_secret"] = secret
    payload = urllib.parse.urlencode(params).encode("utf-8")
    request = urllib.request.Request(
        "https://kauth.kakao.com/oauth/token",
        data=payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = response.read().decode("utf-8")
        return json.loads(raw)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"카카오 토큰 교환 실패: {detail}") from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"카카오 토큰 교환 오류: {exc}") from exc


def _fetch_kakao_userinfo(access_token: str) -> dict[str, Any]:
    request = urllib.request.Request(
        "https://kapi.kakao.com/v2/user/me",
        headers={"Authorization": f"Bearer {access_token}"},
        method="GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = response.read().decode("utf-8")
        return json.loads(raw)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"카카오 사용자 정보 조회 실패: {detail}") from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"카카오 사용자 정보 조회 오류: {exc}") from exc


def _register_oauth_state() -> str:
    """HMAC-signed stateless OAuth state (survives server restarts)."""
    ts = str(int(_utc_now().timestamp()))
    nonce = secrets.token_urlsafe(12)
    payload = f"{ts}.{nonce}"
    sig = hmac_mod.new(_OAUTH_STATE_KEY, payload.encode(), hashlib.sha256).hexdigest()[:32]
    return f"{payload}.{sig}"


def _validate_oauth_state(state: str) -> None:
    """Verify HMAC signature and TTL of OAuth state token."""
    state_norm = (state or "").strip()
    parts = state_norm.split(".")
    if len(parts) != 3:
        raise HTTPException(status_code=400, detail="유효하지 않은 OAuth state입니다.")
    ts_str, nonce, sig = parts
    expected = hmac_mod.new(_OAUTH_STATE_KEY, f"{ts_str}.{nonce}".encode(), hashlib.sha256).hexdigest()[:32]
    if not hmac_mod.compare_digest(sig, expected):
        raise HTTPException(status_code=400, detail="유효하지 않은 OAuth state입니다.")
    try:
        created = datetime.fromtimestamp(int(ts_str), tz=timezone.utc)
    except (ValueError, OSError):
        raise HTTPException(status_code=400, detail="유효하지 않은 OAuth state입니다.")
    if (_utc_now() - created).total_seconds() > OAUTH_STATE_TTL_SECONDS:
        raise HTTPException(status_code=400, detail="만료된 OAuth state입니다.")


def _exchange_google_code_for_token(code: str) -> dict[str, Any]:
    payload = urllib.parse.urlencode(
        {
            "code": code,
            "client_id": _google_client_id(),
            "client_secret": _google_client_secret(),
            "redirect_uri": _google_redirect_uri(),
            "grant_type": "authorization_code",
        }
    ).encode("utf-8")
    request = urllib.request.Request(
        _google_token_url(),
        data=payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = response.read().decode("utf-8")
        return json.loads(raw)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"Google 토큰 교환 실패: {detail}") from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Google 토큰 교환 오류: {exc}") from exc


def _fetch_google_userinfo(access_token: str) -> dict[str, Any]:
    request = urllib.request.Request(
        _google_userinfo_url(),
        headers={"Authorization": f"Bearer {access_token}"},
        method="GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            raw = response.read().decode("utf-8")
        payload = json.loads(raw)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=502, detail=f"Google 사용자 정보 조회 실패: {detail}") from exc
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Google 사용자 정보 조회 오류: {exc}") from exc

    if not isinstance(payload, dict):
        raise HTTPException(status_code=502, detail="Google 사용자 정보 응답 형식 오류")
    return payload




def _sanitize_session_id(session_id: str) -> str:
    normalized = str(session_id or "").strip().lower()
    if not normalized:
        raise HTTPException(status_code=400, detail="session_id가 비어 있습니다.")
    if not re.fullmatch(r"[a-z0-9_-]{8,64}", normalized):
        raise HTTPException(status_code=400, detail="session_id 형식이 올바르지 않습니다.")
    return normalized


def _session_collection_name(session_id: str) -> str:
    safe = re.sub(r"[^a-z0-9_]", "_", session_id)
    return f"web_company_{safe[:48]}"


_SESSION_COMPANY_SCORE_CUTOFF = 0.3
_SESSION_COMPANY_MAX_CHARS = 3000


def _extract_session_company_context(session, rfx_title: str = "") -> str:
    """세션 RAG에서 회사 문서 관련 청크를 추출하여 텍스트로 반환.

    Returns empty string if no company docs in session or extraction fails.
    """
    if not session.rag_engine:
        return ""
    try:
        query = f"{rfx_title} 회사 역량 실적 인력 강점"
        results = session.rag_engine.search(query, top_k=6)
        filtered = [r for r in results if r.score >= _SESSION_COMPANY_SCORE_CUTOFF]
        texts = [r.text for r in filtered if r.text]
        combined = "\n\n".join(texts)[:_SESSION_COMPANY_MAX_CHARS]
        logger.info(
            "session_company_context: raw_hits=%d filtered_hits=%d combined_chars=%d",
            len(results), len(filtered), len(combined),
        )
        return combined
    except Exception as exc:
        logger.debug("Session company context extraction skipped: %s", exc)
        return ""


def _session_upload_dir(session_id: str, bucket: str) -> Path:
    base = ROOT_DIR / "data" / "web_uploads" / session_id / bucket
    base.mkdir(parents=True, exist_ok=True)
    return base


def _cleanup_expired_sessions() -> None:
    expired: list[str] = []
    deadline = _utc_now() - _session_ttl()
    for session_id, session in list(SESSIONS.items()):
        if session.last_used_at < deadline:
            expired.append(session_id)

    for session_id in expired:
        session = SESSIONS.pop(session_id, None)
        if not session:
            continue
        try:
            session.rag_engine.clear_collection()
        except Exception:
            pass
        try:
            session.rfx_rag_engine.clear_collection()
        except Exception:
            pass
        upload_root = ROOT_DIR / "data" / "web_uploads" / session_id
        if upload_root.exists():
            shutil.rmtree(upload_root, ignore_errors=True)


_MAX_SESSIONS = 500


def _get_or_create_session(session_id: str) -> WebRuntimeSession:
    _cleanup_expired_sessions()
    normalized = _sanitize_session_id(session_id)

    existing = SESSIONS.get(normalized)
    if existing:
        existing.last_used_at = _utc_now()
        return existing

    if len(SESSIONS) >= _MAX_SESSIONS:
        raise HTTPException(status_code=503, detail="서버 동시 세션 수를 초과했습니다. 잠시 후 다시 시도해주세요.")

    # RAGEngine은 항상 새로 생성 (in-memory, 재생성 가능)
    rag = RAGEngine(
        persist_directory=str(ROOT_DIR / "data" / "vectordb_web"),
        collection_name=_session_collection_name(normalized),
    )
    rfx_rag = RAGEngine(
        persist_directory=str(ROOT_DIR / "data" / "vectordb_web"),
        collection_name=f"web_rfx_{normalized[:48]}",
    )
    now = _utc_now()
    session = WebRuntimeSession(
        session_id=normalized,
        rag_engine=rag,
        rfx_rag_engine=rfx_rag,
        created_at=now,
        last_used_at=now,
    )

    # Redis에서 분석 결과 복원 시도 (Railway 재시작 후에도 보존)
    analysis_dict = load_analysis_from_redis(normalized)
    if analysis_dict:
        try:
            session.latest_rfx_analysis = restore_analysis_from_dict(analysis_dict)
            logger.info("Restored analysis from Redis for session: %s", normalized)
        except Exception as exc:
            logger.warning("Failed to restore analysis from Redis: %s", exc)

    SESSIONS[normalized] = session
    return session


def _openai_api_key() -> str:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="OPENAI_API_KEY가 설정되지 않았습니다. .env를 확인해주세요.",
        )
    return api_key


def _validate_extension(filename: str) -> None:
    ext = Path(filename or "").suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"지원하지 않는 파일 형식입니다: {ext or '(없음)'}",
        )


_MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50 MB per file


async def _save_upload_file(upload_file: UploadFile, target_dir: Path) -> Path:
    filename = upload_file.filename or f"upload_{uuid.uuid4().hex[:8]}"
    _validate_extension(filename)

    safe_name = re.sub(r"[^0-9A-Za-z._\-가-힣]", "_", filename)[:150]
    dest = target_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:10]}_{safe_name}"

    data = await upload_file.read()
    if len(data) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"파일 크기가 제한(50MB)을 초과합니다: {filename}")
    dest.write_bytes(data)
    return dest


def _serialize_rfx_analysis(analysis: RFxAnalysisResult, rfp_summary: str = "") -> dict[str, Any]:
    return {
        "title": analysis.title,
        "issuing_org": analysis.issuing_org,
        "announcement_number": analysis.announcement_number,
        "deadline": analysis.deadline,
        "project_period": analysis.project_period,
        "budget": analysis.budget,
        "document_type": analysis.document_type,
        "is_rfx_like": analysis.is_rfx_like,
        "document_gate_reason": analysis.document_gate_reason,
        "document_gate_confidence": analysis.document_gate_confidence,
        "requirements": [
            {
                "category": req.category,
                "description": req.description,
                "is_mandatory": req.is_mandatory,
                "detail": req.detail,
            }
            for req in analysis.requirements
        ],
        "evaluation_criteria": [
            {
                "category": item.category,
                "item": item.item,
                "score": item.score,
                "detail": item.detail,
            }
            for item in analysis.evaluation_criteria
        ],
        "required_documents": analysis.required_documents,
        "special_notes": analysis.special_notes,
        "rfp_summary": rfp_summary,
    }


def _serialize_matching_result(result: MatchingResult) -> dict[str, Any]:
    return {
        "overall_score": result.overall_score,
        "recommendation": result.recommendation,
        "summary": result.summary,
        "met_count": result.met_count,
        "partially_met_count": result.partially_met_count,
        "not_met_count": result.not_met_count,
        "unknown_count": result.unknown_count,
        "evaluation": {
            "available": result.evaluation_available,
            "expected_score": result.evaluation_expected_score,
            "total_score": result.evaluation_total_score,
            "technical_expected_score": result.technical_expected_score,
            "price_expected_score": result.price_expected_score,
            "bonus_expected_score": result.bonus_expected_score,
            "notes": result.evaluation_notes,
        },
        "matches": [
            {
                "category": match.requirement.category,
                "description": match.requirement.description,
                "is_mandatory": match.requirement.is_mandatory,
                "status": match.status.value,
                "status_code": {
                    MatchStatus.MET: "MET",
                    MatchStatus.PARTIALLY_MET: "PARTIALLY_MET",
                    MatchStatus.NOT_MET: "NOT_MET",
                    MatchStatus.UNKNOWN: "UNKNOWN",
                }[match.status],
                "confidence": match.confidence,
                "evidence": match.evidence,
                "preparation_guide": match.preparation_guide,
                "source_files": match.source_files,
            }
            for match in result.matches
        ],
        "assistant_opinions": result.assistant_opinions,
        "opinion_mode": result.opinion_mode,
        "report_text": result.to_report(),
    }


def _extract_page_number(source_file: str, metadata: dict[str, Any]) -> int:
    page = int(metadata.get("page_number", -1) or -1)
    if page > 0:
        return page
    match = re.search(r"_p(\d+)$", source_file or "")
    if match:
        return int(match.group(1))
    return -1


def _normalize_ref_text(text: str) -> str:
    normalized = re.sub(r"[^0-9a-zA-Z가-힣\s]", " ", text or "").lower()
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _is_grounded_in_rfx(reference: dict[str, Any], rfx_context: str) -> bool:
    page = int(reference.get("page", 0) or 0)
    if page <= 0:
        return False
    snippet = str(reference.get("text", "")).strip()
    if not snippet:
        return True

    snippet_norm = _normalize_ref_text(snippet)
    rfx_norm = _normalize_ref_text(rfx_context)
    if not snippet_norm or not rfx_norm:
        return False
    if snippet_norm in rfx_norm:
        return True
    keywords = [token for token in snippet_norm.split() if len(token) >= 2][:8]
    hit_count = sum(1 for token in keywords if token in rfx_norm)
    return hit_count >= 1


def _collect_rag_scores(session: WebRuntimeSession, message: str) -> tuple[list[float], list[float]]:
    company_scores: list[float] = []
    rfx_scores: list[float] = []

    try:
        if session.rag_engine.collection.count() > 0:
            company_results = session.rag_engine.search(message, top_k=12)
            company_scores = [float(item.score or 0.0) for item in company_results]
    except Exception:
        company_scores = []

    try:
        if session.rfx_rag_engine.collection.count() > 0:
            rfx_results = session.rfx_rag_engine.search(message, top_k=12)
            rfx_scores = [float(item.score or 0.0) for item in rfx_results]
    except Exception:
        rfx_scores = []

    return company_scores, rfx_scores


def _max_relevance_score(company_scores: list[float], rfx_scores: list[float]) -> float:
    all_scores = company_scores + rfx_scores
    if not all_scores:
        return 0.0
    return max(all_scores)


def _looks_like_gap_question(message: str) -> bool:
    normalized = str(message or "").strip().lower()
    if not normalized:
        return False
    return any(token in normalized for token in GAP_QUERY_KEYWORDS)


def _build_gap_answer_from_latest_matching(
    session: WebRuntimeSession,
    message: str,
) -> tuple[str, list[dict[str, Any]]] | None:
    """
    '부족/미충족' 질의는 최신 매칭 결과를 우선 사용해 결정론적으로 응답한다.
    생성형 답변으로 인한 GO/NO-GO 모순을 방지하기 위한 가드 레이어.
    """
    if not _looks_like_gap_question(message):
        return None

    result = session.latest_matching_result
    if result is None:
        return None

    issue_matches = [
        match
        for match in result.matches
        if match.status in {MatchStatus.NOT_MET, MatchStatus.PARTIALLY_MET, MatchStatus.UNKNOWN}
    ]

    if not issue_matches:
        answer = (
            "최근 분석 결과 기준으로는 **미충족/보완 필요 항목이 없습니다.**\n"
            f"- 종합 적합도: {result.overall_score:.0f}%\n"
            f"- 추천: {result.recommendation}\n\n"
            "이전 분석에서 부족 항목이 있었다면, 그때 사용한 문서(회사 문서 또는 공고 문서)가 "
            "현재 분석 대상과 달랐을 가능성이 큽니다. 동일 문서로 다시 분석해 비교해보세요."
        )
        return answer, []

    lines = [
        "최근 분석 결과 기준 보완 필요 항목입니다.",
        f"- 종합 적합도: {result.overall_score:.0f}%",
        f"- 추천: {result.recommendation}",
        "",
    ]
    for idx, match in enumerate(issue_matches[:6], start=1):
        mandatory = "필수" if match.requirement.is_mandatory else "권장"
        guide = (match.preparation_guide or "증빙 자료 보강이 필요합니다.").strip()
        lines.append(
            f"{idx}. [{mandatory}/{match.status.value}] {match.requirement.description}\n"
            f"   - 대응: {guide}"
        )
    return "\n".join(lines), []


_ORDINAL_MAP: list[tuple[re.Pattern, int]] = [
    (re.compile(r"(?:첫\s*번째|첫째|1\s*번째|첫|제\s*1)"), 0),
    (re.compile(r"(?:두\s*번째|둘째|2\s*번째|제\s*2)"), 1),
    (re.compile(r"(?:세\s*번째|셋째|3\s*번째|제\s*3)"), 2),
    (re.compile(r"(?:네\s*번째|넷째|4\s*번째|제\s*4)"), 3),
    (re.compile(r"(?:다섯\s*번째|5\s*번째|제\s*5)"), 4),
]
_LAST_PAT = re.compile(r"(?:마지막|최근|맨\s*끝)")
_COMPARE_PAT = re.compile(r"(?:비교|가장|중에서|어떤\s*문서|어느\s*문서|모든\s*문서|전체\s*문서|n개)")
_DOC_CONTEXT_PAT = re.compile(r"(?:문서|파일|자료|서류)")


def _resolve_doc_scope(
    message: str,
    explicit_files: list[str] | None,
    company_docs: list[str],
    rfx_docs: list[str],
) -> tuple[list[str] | None, bool]:
    """메시지에서 문서 범위를 결정한다.

    Returns:
        (source_files, is_compare)
        - source_files: 필터할 문서명 리스트 (None이면 전체)
        - is_compare: 비교 모드 여부
    """
    # 1. 프론트엔드 명시 지정 → 최우선
    if explicit_files:
        if "*" in explicit_files:
            return None, True  # 전체 비교
        return explicit_files, len(explicit_files) > 1

    all_docs = company_docs + rfx_docs
    if not all_docs:
        return None, False

    # 2. 비교 키워드 감지
    if _COMPARE_PAT.search(message) and _DOC_CONTEXT_PAT.search(message):
        return None, True

    # 3. 서수 감지
    for pat, idx in _ORDINAL_MAP:
        if pat.search(message) and _DOC_CONTEXT_PAT.search(message):
            if idx < len(all_docs):
                return [all_docs[idx]], False

    # 4. 마지막 문서
    if _LAST_PAT.search(message) and _DOC_CONTEXT_PAT.search(message):
        return [all_docs[-1]], False

    # 5. 파일명 직접 언급
    for doc in all_docs:
        name_no_ext = doc.rsplit(".", 1)[0] if "." in doc else doc
        if name_no_ext in message or doc in message:
            return [doc], False

    return None, False


def _build_chat_context(
    session: WebRuntimeSession,
    message: str,
    source_files: list[str] | None = None,
    is_compare: bool = False,
) -> tuple[str, str, list[dict[str, Any]]]:
    company_context_text = ""
    rfx_context_text = ""
    fallback_refs: list[dict[str, Any]] = []

    # ── 회사 문서 검색 ──
    if session.rag_engine.collection.count() > 0:
        if is_compare:
            for doc_info in session.rag_engine.list_documents():
                sf = doc_info["source_file"]
                results = session.rag_engine.search(message, top_k=6,
                    filter_metadata={"source_file": sf})
                if results:
                    company_context_text += f"\n=== 문서: {sf} ===\n"
                    for result in results:
                        page_num = _extract_page_number(result.source_file, result.metadata)
                        company_context_text += f"[{sf}, 페이지 {page_num}]\n{result.text}\n---\n"
        elif source_files:
            for sf in source_files:
                results = session.rag_engine.search(message, top_k=12,
                    filter_metadata={"source_file": sf})
                for result in results:
                    page_num = _extract_page_number(result.source_file, result.metadata)
                    company_context_text += f"[{sf}, 페이지 {page_num}]\n{result.text}\n---\n"
        else:
            for result in session.rag_engine.search(message, top_k=12):
                page_num = _extract_page_number(result.source_file, result.metadata)
                company_context_text += (
                    f"\n[회사 출처: {result.source_file}, 페이지 {page_num}]\n{result.text}\n---\n"
                )

    # ── RFx 문서 검색 ──
    if session.rfx_rag_engine.collection.count() > 0:
        if is_compare:
            for doc_info in session.rfx_rag_engine.list_documents():
                sf = doc_info["source_file"]
                results = session.rfx_rag_engine.search(message, top_k=6,
                    filter_metadata={"source_file": sf})
                if results:
                    rfx_context_text += f"\n=== 문서: {sf} ===\n"
                    for result in results:
                        page_num = _extract_page_number(result.source_file, result.metadata)
                        rfx_context_text += f"[{sf}, 페이지 {page_num}]\n{result.text}\n---\n"
                        if page_num > 0:
                            fallback_refs.append({"page": page_num, "text": str(result.text or "")[:500]})
        elif source_files:
            for sf in source_files:
                results = session.rfx_rag_engine.search(message, top_k=12,
                    filter_metadata={"source_file": sf})
                for result in results:
                    page_num = _extract_page_number(result.source_file, result.metadata)
                    rfx_context_text += f"[{sf}, 페이지 {page_num}]\n{result.text}\n---\n"
                    if page_num > 0:
                        fallback_refs.append({"page": page_num, "text": str(result.text or "")[:500]})
        else:
            for result in session.rfx_rag_engine.search(message, top_k=12):
                page_num = _extract_page_number(result.source_file, result.metadata)
                rfx_context_text += (
                    f"\n[RFx 출처: {result.source_file}, 페이지 {page_num}]\n{result.text}\n---\n"
                )
                if page_num > 0:
                    fallback_refs.append({"page": page_num, "text": str(result.text or "")[:500]})

    return company_context_text, rfx_context_text, fallback_refs



# ── Tool Use helpers ──

UNSAFE_RESPONSE_TEXT = "해당 요청은 안전 정책상 처리할 수 없습니다."


def _contains_unsafe_keywords(message: str) -> bool:
    """LLM 호출 전 안전 키워드 사전 검사."""
    normalized = " ".join((message or "").strip().lower().split())
    return any(kw in normalized for kw in UNSAFE_KEYWORDS)


def _build_suggested_questions_simple(session: WebRuntimeSession) -> list[str]:
    """Tool Use용 간소화된 추천 질문."""
    if session.latest_matching_result is not None:
        return list(DEFAULT_SUGGESTIONS_WITH_ANALYSIS)
    elif session.rag_engine.collection.count() > 0:
        return list(DEFAULT_SUGGESTIONS_COMPANY_ONLY)
    else:
        return list(DEFAULT_SUGGESTIONS_NO_CONTEXT)


def _write_tool_telemetry(
    *,
    message: str,
    tool_name: str,
    has_context: bool,
    company_scores: list[float],
    rfx_scores: list[float],
) -> None:
    """Tool Use 텔레메트리 (기존 write_router_telemetry 어댑터)."""
    tool_map = {
        "document_qa": RouteIntent.DOMAIN_RFX,
        "general_response": RouteIntent.SMALL_TALK_OFFTOPIC,
        "bid_search": RouteIntent.DOMAIN_RFX,
        "bid_analyze": RouteIntent.DOMAIN_RFX,
        "unsafe_precheck": RouteIntent.UNSAFE,
    }
    intent = tool_map.get(tool_name, RouteIntent.UNKNOWN)
    decision = RouteDecision(
        intent=intent,
        confidence=1.0,
        policy=ChatPolicy.BLOCK_UNSAFE if intent == RouteIntent.UNSAFE else ChatPolicy.ALLOW,
        reason=f"tool_use:{tool_name}",
        source="tool_use",
        llm_called=(tool_name != "unsafe_precheck"),
    )
    rel = max(company_scores + rfx_scores) if (company_scores + rfx_scores) else 0.0
    write_router_telemetry(
        log_path=default_router_log_path(),
        message=message,
        decision=decision,
        company_scores=company_scores,
        rfx_scores=rfx_scores,
        relevance_score=rel,
        min_relevance_score=0.0,
        has_context=has_context,
    )





def _index_rfx_document(session: WebRuntimeSession, file_path: str, original_name: str) -> None:
    parser = DocumentParser()
    parsed = parser.parse(file_path)
    session.rfx_rag_engine.clear_collection()

    if parsed.pages:
        for page_idx, page_text in enumerate(parsed.pages, start=1):
            session.rfx_rag_engine.add_text_directly(
                page_text,
                source_name=f"rfx_{original_name}_p{page_idx}",
                base_metadata={"page_number": page_idx, "type": "rfx_text"},
            )
        return

    if parsed.text:
        session.rfx_rag_engine.add_text_directly(
            parsed.text,
            source_name=f"rfx_{original_name}",
            base_metadata={"page_number": -1, "type": "rfx_text"},
        )


@app.get("/healthz")
def healthz() -> dict[str, str]:
    return {"ok": "true", "time": _utc_now().isoformat()}


@app.get("/api/health")
def health_check() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/debug/env")
@limiter.limit("10/minute")
async def debug_env(request: Request) -> dict[str, Any]:
    """Debug endpoint: check file paths and rag_engine connectivity. 관리자 전용."""
    _require_admin(request)
    import httpx

    rag_health = "unknown"
    rag_error = None
    try:
        fastapi_url = os.environ.get("FASTAPI_URL", "http://localhost:8001")
        async with httpx.AsyncClient(timeout=5) as client:
            resp = await client.get(f"{fastapi_url}/api/debug/env")
            if resp.status_code == 200:
                rag_health = resp.json()
            else:
                rag_error = f"{resp.status_code}: {resp.text[:200]}"
    except Exception as exc:
        rag_error = str(exc)

    return {
        "web_app_cwd": os.getcwd(),
        "fastapi_url": os.environ.get("FASTAPI_URL", "http://localhost:8001"),
        "rag_engine_health": rag_health,
        "rag_engine_error": rag_error,
        "openai_key_set": bool(os.getenv("OPENAI_API_KEY")),
        "port": os.getenv("PORT", "8000"),
    }


@app.get("/api/debug/oauth")
@limiter.limit("10/minute")
def debug_oauth(request: Request) -> dict[str, Any]:
    """OAuth 환경변수 진단 (값은 마스킹). 관리자 전용."""
    _require_admin(request)

    google_client_id = os.getenv("GOOGLE_OAUTH_CLIENT_ID", "").strip()
    google_client_secret = os.getenv("GOOGLE_OAUTH_CLIENT_SECRET", "").strip()
    google_redirect_uri = os.getenv("GOOGLE_OAUTH_REDIRECT_URI", "").strip()
    google_post_login = os.getenv("GOOGLE_OAUTH_POST_LOGIN_URL", "").strip()

    kakao_client_id = os.getenv("KAKAO_CLIENT_ID", "").strip()
    kakao_client_secret = os.getenv("KAKAO_CLIENT_SECRET", "").strip()
    kakao_redirect_uri = os.getenv("KAKAO_REDIRECT_URI", "").strip()

    auth_cookie_secure = os.getenv("AUTH_COOKIE_SECURE", "").strip()

    return {
        "google": {
            "client_id_set": bool(google_client_id),
            "client_id_preview": google_client_id[:12] + "***" if google_client_id else "(empty)",
            "client_secret_set": bool(google_client_secret),
            "redirect_uri": google_redirect_uri or "(not set)",
            "post_login_url": google_post_login or "(not set)",
        },
        "kakao": {
            "client_id_set": bool(kakao_client_id),
            "client_id_preview": kakao_client_id[:8] + "***" if kakao_client_id else "(empty)",
            "client_secret_set": bool(kakao_client_secret),
            "redirect_uri": kakao_redirect_uri or "(not set)",
        },
        "auth_cookie_secure": auth_cookie_secure or "0",
    }


@app.get("/api/debug/smtp")
@limiter.limit("10/minute")
def debug_smtp(request: Request) -> dict[str, Any]:
    """이메일 환경변수 진단 (값은 마스킹). 관리자 전용."""
    _require_admin(request)
    smtp_email = os.getenv("SMTP_EMAIL", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "").strip()
    brevo_key = os.getenv("BREVO_API_KEY", "").strip()
    sender_name = os.getenv("SMTP_SENDER_NAME", "키라봇").strip()
    return {
        "brevo_set": bool(brevo_key),
        "brevo_preview": brevo_key[:8] + "***" if brevo_key else "(empty)",
        "smtp_email_set": bool(smtp_email),
        "smtp_email_preview": smtp_email[:3] + "***" if smtp_email else "(empty)",
        "smtp_password_set": bool(smtp_password),
        "sender_name": sender_name,
        "method": "brevo" if brevo_key else "smtp",
    }


@app.post("/api/debug/smtp-test")
@limiter.limit("10/minute")
def debug_smtp_test(request: Request, payload: dict) -> dict[str, Any]:
    """이메일 테스트 발송 — Brevo 직접 호출로 에러 상세 확인. 관리자 전용."""
    _require_admin(request)
    to = payload.get("to", "").strip()
    if not to:
        return {"ok": False, "error": "to 필요"}

    brevo_key = os.getenv("BREVO_API_KEY", "").strip()
    sender_email = os.getenv("SMTP_EMAIL", "").strip()
    sender_name = os.getenv("SMTP_SENDER_NAME", "키라봇").strip()

    if brevo_key:
        try:
            resp = httpx.post(
                "https://api.brevo.com/v3/smtp/email",
                headers={"api-key": brevo_key, "Content-Type": "application/json"},
                json={
                    "sender": {"email": sender_email or "noreply@kirabot.co.kr", "name": sender_name},
                    "to": [{"email": to}],
                    "subject": "[키라봇] 이메일 테스트",
                    "htmlContent": "<h2>이메일 테스트 성공</h2><p>Brevo API로 발송됨</p>",
                },
                timeout=30,
            )
            return {"ok": resp.status_code in (200, 201), "method": "brevo",
                    "status": resp.status_code, "response": resp.text[:500]}
        except Exception as e:
            return {"ok": False, "method": "brevo", "error": str(e)}
    else:
        try:
            sent = _send_email_smtp(to, "[키라봇] 이메일 테스트", "<h2>SMTP 테스트 성공</h2>")
            return {"ok": sent, "method": "smtp"}
        except Exception as e:
            return {"ok": False, "method": "smtp", "error": str(e)}


@app.get("/")
def landing() -> FileResponse:
    dist_index = FRONTEND_DIST_DIR / "index.html"
    if dist_index.exists():
        return FileResponse(dist_index)
    raise HTTPException(
        status_code=503,
        detail="Frontend dist가 없습니다. 개발 모드는 http://localhost:3000, 배포 모드는 frontend/kirabot 빌드 후 실행하세요.",
    )


@app.get("/index.html")
def landing_html() -> FileResponse:
    return landing()


@app.get("/auth/google/login")
def auth_google_login() -> RedirectResponse:
    state = _register_oauth_state()
    params = {
        "client_id": _google_client_id(),
        "redirect_uri": _google_redirect_uri(),
        "response_type": "code",
        "scope": _google_scopes(),
        "state": state,
        "access_type": "offline",
        "prompt": "consent",
    }
    authorize_url = f"{_google_auth_url()}?{urllib.parse.urlencode(params)}"
    return RedirectResponse(url=authorize_url, status_code=302)


@app.get("/auth/google/callback")
def auth_google_callback(
    request: Request,
    code: str = "",
    state: str = "",
    error: str = "",
) -> RedirectResponse:
    post_url = _post_login_url(request)
    if error:
        target = f"{post_url}?login_error={urllib.parse.quote_plus(error)}"
        return RedirectResponse(url=target, status_code=302)

    _validate_oauth_state(state)
    if not code:
        raise HTTPException(status_code=400, detail="Google authorization code가 없습니다.")

    token_payload = _exchange_google_code_for_token(code)
    access_token = str(token_payload.get("access_token", "")).strip()
    if not access_token:
        raise HTTPException(status_code=502, detail="Google access_token을 받지 못했습니다.")

    userinfo = _fetch_google_userinfo(access_token)
    external_sub = str(userinfo.get("sub", "")).strip()
    email = str(userinfo.get("email", "")).strip().lower()
    display_name = str(userinfo.get("name", "")).strip()
    if not external_sub:
        raise HTTPException(status_code=502, detail="Google 사용자 식별자(sub)가 없습니다.")

    username = upsert_social_user(
        provider="google",
        external_sub=external_sub,
        email=email,
        display_name=display_name,
    )
    ttl_days = max(1, _safe_int_env("AUTH_SESSION_TTL_DAYS", 7))
    session_token = create_user_session(username=username, ttl_days=ttl_days)

    response = RedirectResponse(url=post_url, status_code=302)
    response.set_cookie(
        key=_auth_cookie_name(),
        value=session_token,
        httponly=True,
        secure=_auth_cookie_secure(),
        samesite="lax",
        path=_auth_cookie_path(),
        domain=_auth_cookie_domain(),
        max_age=ttl_days * 24 * 60 * 60,
    )
    return response


# ── Kakao OAuth ──

@app.get("/auth/kakao/login")
def auth_kakao_login() -> RedirectResponse:
    state = _register_oauth_state()
    params = {
        "client_id": _kakao_client_id(),
        "redirect_uri": _kakao_redirect_uri(),
        "response_type": "code",
        "state": state,
    }
    authorize_url = f"https://kauth.kakao.com/oauth/authorize?{urllib.parse.urlencode(params)}"
    return RedirectResponse(url=authorize_url, status_code=302)


@app.get("/auth/kakao/callback")
def auth_kakao_callback(
    request: Request,
    code: str = "",
    state: str = "",
    error: str = "",
) -> RedirectResponse:
    post_url = _post_login_url(request)
    if error:
        target = f"{post_url}?login_error={urllib.parse.quote_plus(error)}"
        return RedirectResponse(url=target, status_code=302)

    _validate_oauth_state(state)
    if not code:
        raise HTTPException(status_code=400, detail="카카오 authorization code가 없습니다.")

    token_payload = _exchange_kakao_code_for_token(code)
    access_token = str(token_payload.get("access_token", "")).strip()
    if not access_token:
        raise HTTPException(status_code=502, detail="카카오 access_token을 받지 못했습니다.")

    userinfo = _fetch_kakao_userinfo(access_token)
    kakao_id = str(userinfo.get("id", "")).strip()
    kakao_account = userinfo.get("kakao_account", {}) or {}
    profile = kakao_account.get("profile", {}) or {}
    email = str(kakao_account.get("email", "")).strip().lower()
    display_name = str(profile.get("nickname", "")).strip()
    avatar_url = str(profile.get("profile_image_url", "")).strip()

    if not kakao_id:
        raise HTTPException(status_code=502, detail="카카오 사용자 식별자(id)가 없습니다.")

    username = upsert_social_user(
        provider="kakao",
        external_sub=kakao_id,
        email=email,
        display_name=display_name,
    )
    ttl_days = max(1, _safe_int_env("AUTH_SESSION_TTL_DAYS", 7))
    session_token = create_user_session(username=username, ttl_days=ttl_days)

    response = RedirectResponse(url=post_url, status_code=302)
    response.set_cookie(
        key=_auth_cookie_name(),
        value=session_token,
        httponly=True,
        secure=_auth_cookie_secure(),
        samesite="lax",
        path=_auth_cookie_path(),
        domain=_auth_cookie_domain(),
        max_age=ttl_days * 24 * 60 * 60,
    )
    return response


@app.get("/auth/me")
def auth_me(request: Request) -> dict[str, Any]:
    session_token = request.cookies.get(_auth_cookie_name(), "")
    username = resolve_user_from_session(session_token)
    if not username:
        return {"ok": True, "authenticated": False, "user": None}

    profile = get_user_profile(username)
    return {
        "ok": True,
        "authenticated": True,
        "user": {
            "id": username,
            "username": username,
            "name": username,
            "email": profile.get("email", ""),
            "provider": profile.get("provider", ""),
            "isAdmin": username in _admin_usernames(),
        },
    }


@app.post("/auth/logout")
def auth_logout(request: Request) -> JSONResponse:
    session_token = request.cookies.get(_auth_cookie_name(), "")
    if session_token:
        invalidate_user_session(session_token)

    response = JSONResponse({"ok": True})
    response.delete_cookie(
        key=_auth_cookie_name(),
        path=_auth_cookie_path(),
        domain=_auth_cookie_domain(),
    )
    return response


@app.post("/api/session")
def create_session() -> dict[str, str]:
    session_id = uuid.uuid4().hex[:24]
    _get_or_create_session(session_id)
    return {"session_id": session_id}


@app.post("/api/session/check")
def session_check(payload: SessionPayload) -> dict[str, Any]:
    """Check if a session exists and has analysis data available."""
    normalized = payload.session_id.strip().lower()[:64]
    session = SESSIONS.get(normalized)
    has_analysis = session is not None and session.latest_rfx_analysis is not None
    return {
        "session_id": normalized,
        "exists": session is not None,
        "has_analysis": has_analysis,
        "analysis_title": session.latest_rfx_analysis.title if has_analysis else None,
    }


@app.post("/api/session/stats")
def session_stats(payload: SessionPayload) -> dict[str, Any]:
    session = _get_or_create_session(payload.session_id)
    stats = session.rag_engine.get_stats()
    return {
        "session_id": session.session_id,
        "company_chunks": stats.get("total_documents", 0),
        "created_at": session.created_at.isoformat(),
        "last_used_at": session.last_used_at.isoformat(),
    }


@app.post("/api/usage/me")
def usage_me(payload: SessionPayload, request: Request) -> dict[str, Any]:
    actor_key, username = _resolve_usage_actor(request, payload.session_id)
    chat_counts = get_actor_usage_counts(actor_key=actor_key, action="chat")
    analyze_counts = get_actor_usage_counts(actor_key=actor_key, action="analyze")
    chat_limits = _usage_limits_for_action("chat")
    analyze_limits = _usage_limits_for_action("analyze")
    return {
        "ok": True,
        "quota_enabled": _quota_enabled(),
        "actor_key": actor_key,
        "username": username,
        "chat": {
            "today_used": chat_counts["today"],
            "month_used": chat_counts["month"],
            "daily_limit": chat_limits[0],
            "monthly_limit": chat_limits[1],
        },
        "analyze": {
            "today_used": analyze_counts["today"],
            "month_used": analyze_counts["month"],
            "daily_limit": analyze_limits[0],
            "monthly_limit": analyze_limits[1],
        },
    }


@app.get("/api/admin/usage")
def admin_usage(request: Request) -> dict[str, Any]:
    token = str(request.cookies.get(_auth_cookie_name(), "") or "")
    username = str(resolve_user_from_session(token) or "")
    if not username or username not in _admin_usernames():
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")
    return {
        "ok": True,
        "username": username,
        "overview": get_usage_overview(),
        "by_actor": list_usage_by_actor(limit=300),
    }


def _require_admin(request: Request) -> str:
    """관리자 인증 공통 헬퍼. 비관리자 → 403."""
    token = str(request.cookies.get(_auth_cookie_name(), "") or "")
    username = str(resolve_user_from_session(token) or "")
    if not username or username not in _admin_usernames():
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")
    return username


@app.get("/api/admin/alerts")
@limiter.limit("10/minute")
def admin_alerts_list(request: Request) -> dict[str, Any]:
    """모든 알림 설정 + 상태 목록 (관리자 전용)."""
    _require_admin(request)

    configs: list[dict[str, Any]] = []
    if ALERT_CONFIG_DIR.exists():
        for f in sorted(ALERT_CONFIG_DIR.glob("*.json")):
            try:
                data = json.loads(f.read_text("utf-8"))
                session_id = f.stem
                state = _get_alert_state(session_id)
                configs.append({
                    "session_id": session_id,
                    "config": data,
                    "state": state,
                })
            except Exception:
                continue

    return {"ok": True, "alerts": configs}


@app.delete("/api/admin/alerts/{config_id}")
@limiter.limit("10/minute")
def admin_alert_delete(config_id: str, request: Request) -> dict[str, Any]:
    """특정 알림 설정 삭제 (관리자 전용)."""
    _require_admin(request)

    if not re.fullmatch(r"[a-zA-Z0-9_\-]{4,128}", config_id):
        raise HTTPException(status_code=400, detail="잘못된 config_id입니다.")

    config_path = ALERT_CONFIG_DIR / f"{config_id}.json"
    if config_path.exists():
        config_path.unlink()

    state_path = ALERT_STATE_DIR / f"{config_id}.json"
    if state_path.exists():
        state_path.unlink()

    return {"ok": True}


@app.post("/api/admin/alerts/{config_id}/send-now")
@limiter.limit("10/minute")
async def admin_alert_send_now(config_id: str, request: Request) -> dict[str, Any]:
    """특정 알림 즉시 발송 (관리자 전용)."""
    _require_admin(request)

    if not re.fullmatch(r"[a-zA-Z0-9_\-]{4,128}", config_id):
        raise HTTPException(status_code=400, detail="잘못된 config_id입니다.")

    config_path = ALERT_CONFIG_DIR / f"{config_id}.json"
    if not config_path.exists():
        raise HTTPException(status_code=404, detail="알림 설정을 찾을 수 없습니다.")

    config = json.loads(config_path.read_text("utf-8"))
    if not config.get("email"):
        raise HTTPException(status_code=400, detail="수신 이메일이 설정되지 않았습니다.")

    result = await _execute_alert_send(config, label=f"admin-send:{config_id}", session_id=config_id)
    return {"ok": True, **result}


@app.post("/api/company/upload")
async def upload_company_documents(
    session_id: str = Form(...),
    files: list[UploadFile] = File(...),
) -> dict[str, Any]:
    if not files:
        raise HTTPException(status_code=400, detail="업로드 파일이 없습니다.")

    session = _get_or_create_session(session_id)
    upload_dir = _session_upload_dir(session.session_id, "company")

    uploaded_names: list[str] = []
    file_urls: list[str] = []
    total_chunks = 0
    for file in files:
        saved_path = await _save_upload_file(file, upload_dir)
        uploaded_names.append(file.filename or saved_path.name)
        file_urls.append(f"/api/files/{session.session_id}/company/{saved_path.name}")
        total_chunks += await asyncio.to_thread(session.rag_engine.add_document, str(saved_path))

    stats = session.rag_engine.get_stats()
    return {
        "ok": True,
        "uploaded_files": uploaded_names,
        "fileUrls": file_urls,
        "added_chunks": total_chunks,
        "company_chunks": stats.get("total_documents", 0),
        "documents": session.rag_engine.list_documents(),
    }


@app.post("/api/company/text")
async def upload_company_text(payload: dict) -> dict[str, Any]:
    """텍스트 기반 회사 정보를 가상 문서로 RAG 엔진에 저장."""
    session_id = payload.get("session_id", "").strip()
    text = payload.get("text", "").strip()
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id가 필요합니다.")
    if not text or len(text) < 10:
        raise HTTPException(status_code=400, detail="회사 정보를 10자 이상 입력해주세요.")

    session = _get_or_create_session(session_id)
    upload_dir = _session_upload_dir(session.session_id, "company")

    # 텍스트를 company_info.txt로 저장
    text_path = upload_dir / "company_info.txt"
    # 기존 텍스트 파일이 있으면 내용 추가
    if text_path.exists():
        prev = text_path.read_text("utf-8")
        text_path.write_text(prev + "\n\n" + text, encoding="utf-8")
    else:
        text_path.write_text(text, encoding="utf-8")

    added_chunks = session.rag_engine.add_document(str(text_path))
    stats = session.rag_engine.get_stats()

    return {
        "ok": True,
        "added_chunks": added_chunks,
        "company_chunks": stats.get("total_documents", 0),
        "documents": session.rag_engine.list_documents(),
        "fileUrls": [f"/api/files/{session.session_id}/company/company_info.txt"],
        "uploaded_files": ["company_info.txt"],
    }


@app.get("/api/company/list")
def list_company_documents(session_id: str) -> dict[str, Any]:
    session = _get_or_create_session(session_id)
    docs = session.rag_engine.list_documents()

    # 디스크 파일과 매칭하여 URL 생성
    upload_dir = ROOT_DIR / "data" / "web_uploads" / session.session_id / "company"
    for doc in docs:
        matched_file = None
        if upload_dir.exists():
            for f in upload_dir.iterdir():
                if f.name.endswith(doc["source_file"]) or doc["source_file"] in f.name:
                    matched_file = f.name
                    break
        doc["url"] = f"/api/files/{session.session_id}/company/{matched_file}" if matched_file else ""

    return {
        "ok": True,
        "documents": docs,
        "total_chunks": sum(d["chunks"] for d in docs),
    }


@app.post("/api/company/delete")
def delete_company_document(payload: CompanyDeletePayload) -> dict[str, Any]:
    session = _get_or_create_session(payload.session_id)
    deleted = session.rag_engine.delete_document(payload.source_file)

    if deleted == 0:
        raise HTTPException(status_code=404, detail="해당 문서를 찾을 수 없습니다.")

    # 디스크에서도 삭제
    upload_dir = ROOT_DIR / "data" / "web_uploads" / session.session_id / "company"
    if upload_dir.exists():
        for f in upload_dir.iterdir():
            if f.name.endswith(payload.source_file) or payload.source_file in f.name:
                f.unlink(missing_ok=True)
                break

    # 매칭 결과 무효화
    session.latest_matching_result = None

    remaining = session.rag_engine.get_stats().get("total_documents", 0)
    return {
        "ok": True,
        "deleted_chunks": deleted,
        "remaining_chunks": remaining,
    }


@app.post("/api/company/clear")
def clear_company_documents(payload: SessionPayload) -> dict[str, Any]:
    session = _get_or_create_session(payload.session_id)
    session.rag_engine.clear_collection()

    upload_dir = ROOT_DIR / "data" / "web_uploads" / session.session_id / "company"
    if upload_dir.exists():
        shutil.rmtree(upload_dir, ignore_errors=True)

    return {"ok": True, "company_chunks": 0}


@app.post("/api/analyze/upload")
@limiter.limit("10/minute")
async def analyze_uploaded_document(
    request: Request,
    session_id: str = Form(...),
    files: list[UploadFile] = File(...),
) -> dict[str, Any]:
    from services.web_app.structured_logger import get_structured_logger
    slog = get_structured_logger("web_app.analyze")
    slog.set_request_id(getattr(request.state, "request_id", None))

    session = _get_or_create_session(session_id)
    company_chunk_count = _inject_company_profile_if_needed(session, request)

    slog.info(
        "analysis_started",
        session_id=session_id,
        filename=files[0].filename,
        file_count=len(files),
        file_size=files[0].size if hasattr(files[0], 'size') else None,
        company_chunks=company_chunk_count,
    )

    quota = _enforce_quota_or_raise(
        request=request,
        session_id=session.session_id,
        action="analyze",
        metadata={
            "path": "/api/analyze/upload",
            "filename": str(files[0].filename or ""),
        },
    )

    api_key = _openai_api_key()
    upload_dir = _session_upload_dir(session.session_id, "target")

    # Save all uploaded files
    saved_paths: list[Path] = []
    for f in files:
        saved_paths.append(await _save_upload_file(f, upload_dir))

    primary_path = saved_paths[0]
    primary_filename = files[0].filename or primary_path.name

    analyzer = RFxAnalyzer(
        api_key=api_key,
        model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
    )

    try:
        # Increase timeout for large PDFs (58+ pages can take 2-3 minutes)
        analysis = await asyncio.wait_for(
            asyncio.to_thread(analyzer.analyze, str(primary_path)),
            timeout=240.0  # 4 minutes (프론트엔드 180초보다 여유있게)
        )
        slog.info("analysis_completed", session_id=session_id, filename=primary_filename)
    except asyncio.TimeoutError:
        slog.error("analysis_timeout", session_id=session_id, filename=primary_filename, timeout_sec=240)
        raise HTTPException(
            status_code=504,
            detail="문서 분석 시간이 초과되었습니다 (4분). 페이지 수가 너무 많거나 복잡한 문서일 수 있습니다."
        )
    except ImportError as exc:
        slog.error("analysis_import_error", session_id=session_id, error=str(exc))
        raise HTTPException(status_code=503, detail="서비스를 일시적으로 사용할 수 없습니다.") from exc
    except ValueError as exc:
        slog.error("analysis_value_error", session_id=session_id, error=str(exc))
        raise HTTPException(status_code=400, detail=f"문서 분석 실패: {exc}") from exc
    except Exception as exc:
        slog.error("analysis_failed", session_id=session_id, error=str(exc), traceback=traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"문서 분석 중 오류: {exc}") from exc

    # RFP 요약 + 매칭을 동시 실행 (둘 다 분석 결과만 필요, 서로 독립적)
    rfp_summary = ""
    matching = None

    async def _gen_summary() -> str:
        if not analysis.raw_text:
            return ""
        try:
            return await asyncio.to_thread(analyzer.generate_rfp_summary, analysis.raw_text)
        except Exception as exc:
            logger.warning("RFP summary generation failed: %s", exc)
            return ""

    async def _run_matching():
        if company_chunk_count <= 0:
            return None
        try:
            matcher = QualificationMatcher(
                rag_engine=session.rag_engine,
                api_key=api_key,
                model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            )
            return await asyncio.to_thread(matcher.match, analysis)
        except Exception as exc:
            logger.warning("Matching failed: %s", exc)
            return None

    rfp_summary, matching = await asyncio.gather(_gen_summary(), _run_matching())

    # Index all uploaded files into RAG for Q&A
    for i, sp in enumerate(saved_paths):
        original_name = files[i].filename or "target"
        await asyncio.to_thread(
            _index_rfx_document, session=session, file_path=str(sp), original_name=original_name,
        )

    session.latest_rfx_analysis = analysis
    session.latest_matching_result = matching
    session.latest_document_name = primary_filename

    # Redis에 분석 결과 저장 (Railway 재시작 대비)
    try:
        save_analysis_to_redis(session.session_id, asdict(analysis))
    except Exception as exc:
        logger.warning("Failed to save analysis to Redis: %s", exc)

    # DB write-through: AnalysisSnapshot 저장 (Task 9)
    if _BID_DB_ENABLED:
        try:
            from services.web_app.api.adapter import SessionAdapter
            from services.web_app.api.analysis_serializer import serialize_analysis_for_db
            from services.web_app.db.engine import get_async_session_factory

            async_session_factory = get_async_session_factory()
            async with async_session_factory() as db:
                adapter = SessionAdapter(db)
                _, username = _resolve_usage_actor(request, session.session_id)
                project = await adapter.get_or_create_project(
                    session_id=session.session_id,
                    username=username,
                    title=primary_filename,
                )
                await adapter.save_analysis(
                    project_id=project.id,
                    org_id=project.org_id,
                    analysis_json=serialize_analysis_for_db(analysis),
                    summary_md=rfp_summary or None,
                    go_nogo_json=asdict(matching) if matching else None,
                    username=username,
                )
                slog.info("analysis_persisted_to_db", session_id=session_id, project_id=project.id)
        except Exception as exc:
            # DB 저장 실패해도 API 응답은 유지 (로그만)
            logger.warning("Failed to persist analysis to DB (session=%s): %s", session.session_id, exc)

    file_urls = [f"/api/files/{session.session_id}/target/{sp.name}" for sp in saved_paths]
    filenames = [f.filename or sp.name for f, sp in zip(files, saved_paths)]

    serialized = _serialize_rfx_analysis(analysis, rfp_summary=rfp_summary)

    return {
        "ok": True,
        "filename": primary_filename,  # backward compat
        "filenames": filenames,  # NEW: all filenames
        "session_id": session.session_id,
        "company_chunks": company_chunk_count,
        "quota": quota,
        "analysis": serialized,
        "matching": _serialize_matching_result(matching) if matching else None,
        "fileUrl": file_urls[0],  # backward compat
        "fileUrls": file_urls,  # NEW: all file URLs
    }


@app.post("/api/analyze/text")
@limiter.limit("10/minute")
async def analyze_text(payload: AnalyzeTextPayload, request: Request) -> dict[str, Any]:
    """Analyze text input (async converted from sync for Task 9 DB write-through)."""
    session = _get_or_create_session(payload.session_id)
    company_chunk_count = _inject_company_profile_if_needed(session, request)

    quota = _enforce_quota_or_raise(
        request=request,
        session_id=session.session_id,
        action="analyze",
        metadata={"path": "/api/analyze/text", "source": "direct_text"},
    )

    api_key = _openai_api_key()
    analyzer = RFxAnalyzer(
        api_key=api_key,
        model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
    )

    # Convert blocking calls to async
    analysis = await asyncio.to_thread(analyzer.analyze_text, payload.document_text)

    if company_chunk_count > 0:
        matcher = QualificationMatcher(
            rag_engine=session.rag_engine,
            api_key=api_key,
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
        )
        matching = await asyncio.to_thread(matcher.match, analysis)
    else:
        matching = None

    # RAG indexing (blocking, but fast)
    await asyncio.to_thread(session.rfx_rag_engine.clear_collection)
    await asyncio.to_thread(
        session.rfx_rag_engine.add_text_directly,
        payload.document_text,
        source_name="rfx_direct_input",
        base_metadata={"page_number": -1, "type": "rfx_text"},
    )

    session.latest_rfx_analysis = analysis
    session.latest_matching_result = matching
    session.latest_document_name = "direct_input.txt"

    # Redis에 분석 결과 저장
    try:
        save_analysis_to_redis(session.session_id, asdict(analysis))
    except Exception as exc:
        logger.warning("Failed to save analysis to Redis: %s", exc)

    # DB write-through: AnalysisSnapshot 저장 (Task 9)
    if _BID_DB_ENABLED:
        try:
            from services.web_app.api.adapter import SessionAdapter
            from services.web_app.api.analysis_serializer import serialize_analysis_for_db
            from services.web_app.db.engine import get_async_session_factory

            async_session_factory = get_async_session_factory()
            async with async_session_factory() as db:
                adapter = SessionAdapter(db)
                _, username = _resolve_usage_actor(request, session.session_id)
                project = await adapter.get_or_create_project(
                    session_id=session.session_id,
                    username=username,
                    title="직접 입력 텍스트",
                )
                await adapter.save_analysis(
                    project_id=project.id,
                    org_id=project.org_id,
                    analysis_json=serialize_analysis_for_db(analysis),
                    summary_md=None,
                    go_nogo_json=asdict(matching) if matching else None,
                    username=username,
                )
        except Exception as exc:
            # DB 저장 실패해도 API 응답은 유지 (로그만)
            logger.warning("Failed to persist analysis to DB (session=%s): %s", session.session_id, exc)

    return {
        "ok": True,
        "session_id": session.session_id,
        "company_chunks": company_chunk_count,
        "quota": quota,
        "analysis": _serialize_rfx_analysis(analysis),
        "matching": _serialize_matching_result(matching) if matching else None,
    }


class RematchPayload(BaseModel):
    session_id: str


@app.post("/api/rematch")
def rematch_with_company_docs(payload: RematchPayload, request: Request) -> dict[str, Any]:
    """회사 문서 등록 후 기존 분석 결과에 대해 매칭만 재실행."""
    session = _get_or_create_session(payload.session_id)
    if session.latest_rfx_analysis is None:
        raise HTTPException(status_code=400, detail="재매칭할 분석 결과가 없습니다.")

    company_chunk_count = _inject_company_profile_if_needed(session, request)
    if company_chunk_count <= 0:
        raise HTTPException(status_code=400, detail="등록된 회사 문서가 없습니다.")

    api_key = _openai_api_key()
    matcher = QualificationMatcher(
        rag_engine=session.rag_engine,
        api_key=api_key,
        model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
    )
    matching = matcher.match(session.latest_rfx_analysis)
    session.latest_matching_result = matching

    return {
        "ok": True,
        "session_id": session.session_id,
        "company_chunks": company_chunk_count,
        "analysis": _serialize_rfx_analysis(session.latest_rfx_analysis),
        "matching": _serialize_matching_result(matching) if matching else None,
        "filename": session.latest_document_name,
    }


@app.post("/api/chat")
@limiter.limit("30/minute")
def chat_with_references(payload: ChatPayload, request: Request) -> dict[str, Any]:
    session = _get_or_create_session(payload.session_id)
    # 회사 프로필 문서 자동 주입 (RAG 검색에서 회사 정보 활용)
    _inject_company_profile_if_needed(session, request)
    message = str(payload.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="message가 비어 있습니다.")

    quota = _enforce_quota_or_raise(
        request=request,
        session_id=session.session_id,
        action="chat",
        metadata={"path": "/api/chat"},
    )

    # ── Step 0: UNSAFE 키워드 사전 검사 (LLM 호출 전 비용 절약) ──
    if _contains_unsafe_keywords(message):
        _write_tool_telemetry(
            message=message,
            tool_name="unsafe_precheck",
            has_context=False,
            company_scores=[],
            rfx_scores=[],
        )
        return {
            "ok": True,
            "blocked": True,
            "policy": "BLOCK_UNSAFE",
            "intent": "UNSAFE",
            "reason": "안전 키워드 감지",
            "answer": UNSAFE_RESPONSE_TEXT,
            "references": [],
            "suggested_questions": _build_suggested_questions_simple(session),
            "quota": quota,
        }

    # ── Step 1: 결정론적 갭 답변 가드 (매칭 결과 모순 방지) ──
    deterministic_gap = _build_gap_answer_from_latest_matching(
        session=session, message=message,
    )
    if deterministic_gap is not None:
        answer, references = deterministic_gap
        return {
            "ok": True,
            "blocked": False,
            "policy": "ALLOW",
            "intent": "document_qa",
            "reason": "latest_matching_guard",
            "answer": answer,
            "references": references,
            "suggested_questions": _build_suggested_questions_simple(session),
            "quota": quota,
        }

    # ── Step 2: RAG 컨텍스트 구축 ──
    api_key = _openai_api_key()
    company_docs = [d["source_file"] for d in session.rag_engine.list_documents()]
    rfx_docs = [d["source_file"] for d in session.rfx_rag_engine.list_documents()]
    scope_files, is_compare = _resolve_doc_scope(
        message, payload.source_files, company_docs, rfx_docs
    )
    company_context_text, rfx_context_text, fallback_refs = _build_chat_context(
        session, message, source_files=scope_files, is_compare=is_compare,
    )
    company_scores, rfx_scores = _collect_rag_scores(session, message)
    has_context = bool(
        (session.rag_engine.collection.count() > 0)
        or (session.rfx_rag_engine.collection.count() > 0)
    )

    # ── Step 3: ReAct 루프 (최대 3턴, 재검색 가능) ──
    try:
        tool_name, answer, references = react_chat_loop(
            api_key=api_key,
            message=message,
            company_context_text=company_context_text,
            rfx_context_text=rfx_context_text,
            session=session,
        )
    except Exception as exc:
        logger.error("Tool Use LLM 호출 실패: %s", exc)
        return {
            "ok": True,
            "blocked": False,
            "policy": "ALLOW",
            "intent": "error_fallback",
            "reason": "llm_call_failed",
            "answer": "죄송합니다, 답변 생성 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.",
            "references": [],
            "suggested_questions": _build_suggested_questions_simple(session),
            "quota": quota,
        }

    # ── Step 4: document_qa 참조 그라운딩 ──
    if tool_name == "document_qa":
        references = [
            r for r in references if _is_grounded_in_rfx(r, rfx_context_text)
        ]
        if not references and fallback_refs:
            references = fallback_refs[:3]

    # ── Step 5: 텔레메트리 ──
    _write_tool_telemetry(
        message=message,
        tool_name=tool_name,
        has_context=has_context,
        company_scores=company_scores,
        rfx_scores=rfx_scores,
    )

    return {
        "ok": True,
        "blocked": False,
        "policy": "ALLOW",
        "intent": tool_name,
        "reason": f"tool_use:{tool_name}",
        "answer": answer,
        "references": references,
        "suggested_questions": _build_suggested_questions_simple(session),
        "quota": quota,
        "scoped_to": scope_files,
    }



# ── Phase 1: 공고 검색 (나라장터 Open API) ──


@app.post("/api/bids/search")
@limiter.limit("20/minute")
async def api_bids_search(payload: BidSearchPayload, request: Request) -> dict[str, Any]:
    """나라장터 공고 검색."""
    kw = payload.keywords
    if isinstance(kw, list):
        keywords_str = " ".join(kw)
    else:
        keywords_str = str(kw or "").strip()

    try:
        result = await nara_search_bids(
            keywords=keywords_str,
            category=payload.category,
            region=payload.region,
            min_amt=payload.minAmt,
            max_amt=payload.maxAmt,
            period=payload.period,
            exclude_expired=payload.excludeExpired,
            page=payload.page,
            page_size=payload.pageSize,
        )
    except ValueError as exc:
        logger.error("bid search ValueError: %s", exc)
        raise HTTPException(status_code=503, detail="서비스를 일시적으로 사용할 수 없습니다.") from exc
    except Exception as exc:
        logger.error("bid search failed: %s", exc)
        raise HTTPException(status_code=502, detail="나라장터 API 호출 실패") from exc
    return result


@app.get("/api/bids/{bid_ntce_no}/attachments")
async def api_bid_attachments(bid_ntce_no: str, bid_ntce_ord: str = "00") -> dict[str, Any]:
    """공고 첨부파일 목록 조회."""
    try:
        attachments = await nara_get_bid_attachments(bid_ntce_no, bid_ntce_ord)
    except ValueError as exc:
        logger.error("bid attachments ValueError: %s", exc)
        raise HTTPException(status_code=503, detail="서비스를 일시적으로 사용할 수 없습니다.") from exc
    except Exception as exc:
        logger.error("bid attachments failed: %s", exc)
        raise HTTPException(status_code=502, detail="첨부파일 조회 실패") from exc
    return {"attachments": attachments}


# ── Phase 2: 파일 서빙 (문서 미리보기) ──


@app.get("/api/files/{session_id}/{bucket}/{filename}")
def serve_uploaded_file(session_id: str, bucket: str, filename: str) -> FileResponse:
    """업로드된 파일 서빙 (PDF 미리보기 등)."""
    session_id = _sanitize_session_id(session_id)
    if bucket not in ("company", "target"):
        raise HTTPException(status_code=400, detail="유효하지 않은 버킷입니다.")

    # Path traversal 방어: Path.name으로 디렉토리 컴포넌트 제거 + resolve 검증
    safe_filename = Path(filename).name
    if not safe_filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="유효하지 않은 파일명입니다.")

    file_path = ROOT_DIR / "data" / "web_uploads" / session_id / bucket / safe_filename
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")

    # 파일 경로가 upload 디렉토리 내에 있는지 재검증
    try:
        file_path.resolve().relative_to((ROOT_DIR / "data" / "web_uploads").resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="접근이 허용되지 않은 경로입니다.")

    from urllib.parse import quote
    encoded_name = quote(safe_filename)
    return FileResponse(
        path=str(file_path),
        filename=safe_filename,
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{encoded_name}"},
    )


@app.get("/api/preview/text")
def serve_file_text_preview(session_id: str, bucket: str, filename: str) -> dict[str, Any]:
    """HWP/DOCX 등 비PDF 문서의 텍스트 추출 (미리보기용).

    Query params: ?session_id=...&bucket=...&filename=...
    """
    session_id = _sanitize_session_id(session_id)
    if bucket not in ("company", "target"):
        raise HTTPException(status_code=400, detail="유효하지 않은 버킷입니다.")

    safe_filename = Path(filename).name
    if not safe_filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="유효하지 않은 파일명입니다.")

    file_path = ROOT_DIR / "data" / "web_uploads" / session_id / bucket / safe_filename
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")

    try:
        file_path.resolve().relative_to((ROOT_DIR / "data" / "web_uploads").resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="접근이 허용되지 않은 경로입니다.")

    try:
        from document_parser import DocumentParser
        parser = DocumentParser()
        parsed = parser.parse(str(file_path))
        # parsed.pages is list[str] — each element is a page's text
        if parsed.pages:
            pages = [{"page_number": i + 1, "text": text.strip()} for i, text in enumerate(parsed.pages)]
        else:
            pages = [{"page_number": 1, "text": parsed.text.strip()}]
        return {"fileName": safe_filename, "totalPages": len(pages), "pages": pages}
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"텍스트 추출 실패: {exc}") from exc


# ── Phase 3: 공고 첨부파일 자동분석 ──


@app.post("/api/bids/analyze")
@limiter.limit("10/minute")
async def analyze_bid_from_nara(payload: BidAnalyzePayload, request: Request) -> dict[str, Any]:
    """나라장터 공고 첨부파일 자동 다운로드 + 분석."""
    session = _get_or_create_session(payload.session_id)
    company_chunk_count = _inject_company_profile_if_needed(session, request)

    quota = _enforce_quota_or_raise(
        request=request,
        session_id=session.session_id,
        action="analyze",
        metadata={"path": "/api/bids/analyze", "bid_ntce_no": payload.bid_ntce_no},
    )

    # 1. 첨부파일 3단계 폴백: 상세 API → e발주 API → 수동 업로드
    attachments: list[dict[str, Any]] = []
    attachment_error: str | None = None

    # Tier 1: 상세 API (ntceSpecFileNm1~10 — 공고서 원본/변환본)
    try:
        attachments = await nara_get_bid_detail_attachments(
            payload.bid_ntce_no, payload.bid_ntce_ord, payload.category,
        )
    except Exception as exc:
        logger.info("상세 API 폴백 (bid=%s): %s", payload.bid_ntce_no, exc)

    # Tier 2: e발주 API (제안요청서, 기타문서)
    if not attachments:
        try:
            attachments = await nara_get_bid_attachments(payload.bid_ntce_no, payload.bid_ntce_ord)
        except ValueError as exc:
            attachment_error = str(exc)
        except Exception as exc:
            attachment_error = f"첨부파일 조회 실패: {exc}"
            logger.warning("첨부파일 API 오류 (bid=%s): %s", payload.bid_ntce_no, exc)

    # Tier 3: 수동 업로드 폴백
    if not attachments:
        raise HTTPException(
            status_code=422,
            detail=json.dumps({
                "code": "attachment_unavailable",
                "message": "공고 첨부파일을 자동으로 가져올 수 없습니다. 나라장터에서 직접 다운로드 후 업로드해주세요.",
                "reason": attachment_error or "첨부파일 정보 없음",
                "bidNtceNo": payload.bid_ntce_no,
                "bidNtceOrd": payload.bid_ntce_ord,
            }, ensure_ascii=False),
        )

    # 2. 최적 첨부파일 선택
    best = pick_best_attachment(attachments)
    if not best:
        raise HTTPException(status_code=404, detail="분석 가능한 첨부파일이 없습니다.")

    # 3. 다운로드
    target_dir = _session_upload_dir(session.session_id, "target")
    try:
        local_path = await nara_download_attachment(best["fileUrl"], str(target_dir), fallback_name=best["fileNm"])
    except Exception as exc:
        logger.warning("첨부파일 다운로드 실패 (bid=%s): %s", payload.bid_ntce_no, exc)
        raise HTTPException(
            status_code=422,
            detail=json.dumps({
                "code": "attachment_unavailable",
                "message": "공고 첨부파일 서버에서 자동 다운로드하지 못했습니다. 파일을 직접 업로드하면 같은 분석을 계속 진행할 수 있습니다.",
                "reason": f"다운로드 실패: {exc}",
                "bidNtceNo": payload.bid_ntce_no,
                "bidNtceOrd": payload.bid_ntce_ord,
            }, ensure_ascii=False),
        ) from exc

    # 4. 분석
    api_key = _openai_api_key()
    analyzer = RFxAnalyzer(
        api_key=api_key,
        model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
    )
    matcher = QualificationMatcher(
        rag_engine=session.rag_engine,
        api_key=api_key,
        model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
    )

    try:
        analysis = await asyncio.to_thread(analyzer.analyze, local_path)
    except ImportError as exc:
        logger.error("bid analysis ImportError (bid=%s): %s", payload.bid_ntce_no, exc)
        raise HTTPException(status_code=503, detail="서비스를 일시적으로 사용할 수 없습니다.") from exc
    except ValueError as exc:
        logger.warning("bid analysis ValueError (bid=%s): %s", payload.bid_ntce_no, exc)
        raise HTTPException(status_code=400, detail="문서 분석 실패") from exc
    except Exception as exc:
        logger.warning("Bid analysis failed (bid=%s): %s", payload.bid_ntce_no, exc)
        raise HTTPException(status_code=500, detail="문서 분석 중 오류가 발생했습니다.") from exc

    # RFP 요약 + 매칭을 동시 실행 (둘 다 분석 결과만 필요, 서로 독립적)
    async def _gen_bid_summary() -> str:
        if not analysis.raw_text:
            return ""
        try:
            return await asyncio.to_thread(analyzer.generate_rfp_summary, analysis.raw_text)
        except Exception as exc:
            logger.warning("RFP summary failed (bid=%s): %s", payload.bid_ntce_no, exc)
            return ""

    async def _run_bid_matching():
        if company_chunk_count <= 0:
            return None
        try:
            return await asyncio.to_thread(matcher.match, analysis)
        except Exception as exc:
            logger.warning("Matching failed (bid=%s): %s", payload.bid_ntce_no, exc)
            return None

    rfp_summary, matching = await asyncio.gather(_gen_bid_summary(), _run_bid_matching())

    await asyncio.to_thread(
        _index_rfx_document, session=session, file_path=local_path, original_name=best["fileNm"],
    )
    session.latest_rfx_analysis = analysis
    session.latest_matching_result = matching
    session.latest_document_name = best["fileNm"]

    # Redis에 분석 결과 저장
    try:
        save_analysis_to_redis(session.session_id, asdict(analysis))
    except Exception as exc:
        logger.warning("Failed to save analysis to Redis: %s", exc)

    # DB write-through: AnalysisSnapshot 저장 (Task 9)
    if _BID_DB_ENABLED:
        try:
            from services.web_app.api.adapter import SessionAdapter
            from services.web_app.api.analysis_serializer import serialize_analysis_for_db
            from services.web_app.db.engine import get_async_session_factory

            async_session_factory = get_async_session_factory()
            async with async_session_factory() as db:
                adapter = SessionAdapter(db)
                _, username = _resolve_usage_actor(request, session.session_id)
                project = await adapter.get_or_create_project(
                    session_id=session.session_id,
                    username=username,
                    title=best["fileNm"],  # 첨부파일명을 프로젝트 타이틀로 사용
                )
                await adapter.save_analysis(
                    project_id=project.id,
                    org_id=project.org_id,
                    analysis_json=serialize_analysis_for_db(analysis),
                    summary_md=rfp_summary or None,
                    go_nogo_json=asdict(matching) if matching else None,
                    username=username,
                )
                slog.info("analysis_persisted_to_db", session_id=session.session_id, project_id=project.id)
        except Exception as exc:
            # DB 저장 실패해도 API 응답은 유지 (로그만)
            logger.warning("Failed to persist analysis to DB (session=%s): %s", session.session_id, exc)

    # 5. fileUrl 생성
    local_filename = Path(local_path).name
    file_url = f"/api/files/{session.session_id}/target/{local_filename}"

    serialized = _serialize_rfx_analysis(analysis, rfp_summary=rfp_summary)

    return {
        "ok": True,
        "filename": best["fileNm"],
        "session_id": session.session_id,
        "company_chunks": company_chunk_count,
        "quota": quota,
        "analysis": serialized,
        "matching": _serialize_matching_result(matching) if matching else None,
        "fileUrl": file_url,
    }


# ── 일괄 평가 (레거시) ──


@app.post("/api/bids/evaluate-batch")
@limiter.limit("5/minute")
async def api_bids_evaluate_batch(payload: BidEvaluateBatchPayload, request: Request) -> dict[str, Any]:
    """선택된 공고들을 순차 다운로드 + 분석하여 일괄 평가."""
    session = _get_or_create_session(payload.session_id)
    company_chunk_count = _inject_company_profile_if_needed(session, request)

    api_key = _openai_api_key()
    sem = asyncio.Semaphore(3)

    async def _evaluate_one(bid_ntce_no: str) -> dict[str, Any]:
        job: dict[str, Any] = {
            "id": bid_ntce_no,
            "bidNoticeId": bid_ntce_no,
            "isEligible": None,
            "evaluationReason": "",
            "actionPlan": None,
            "bidNotice": {"id": bid_ntce_no, "title": bid_ntce_no, "region": None, "deadlineAt": None, "url": None},
        }

        async with sem:
            try:
                # 1. 첨부파일 3단계 폴백: 상세 API → e발주 API
                attachments = await nara_get_bid_detail_attachments(bid_ntce_no, "00")
                if not attachments:
                    attachments = await nara_get_bid_attachments(bid_ntce_no, "00")
                best = pick_best_attachment(attachments) if attachments else None
                if not best:
                    job["evaluationReason"] = "첨부파일이 없어 분석 불가"
                    return job

                target_dir = _session_upload_dir(session.session_id, "target")
                local_path = await nara_download_attachment(best["fileUrl"], str(target_dir), fallback_name=best["fileNm"])

                # 2. 분석
                analyzer = RFxAnalyzer(api_key=api_key, model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"))
                analysis = await asyncio.to_thread(analyzer.analyze, local_path)

                matching = None
                if company_chunk_count > 0:
                    matcher = QualificationMatcher(
                        rag_engine=session.rag_engine,
                        api_key=api_key,
                        model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                    )
                    matching = await asyncio.to_thread(matcher.match, analysis)

                job["bidNotice"]["title"] = analysis.title or bid_ntce_no
                job["isEligible"] = matching.recommendation == "GO" if matching else None
                job["evaluationReason"] = matching.summary if matching else "자격요건 추출 완료 (회사 문서 미등록으로 매칭 미수행)"
                if matching and matching.gaps:
                    job["actionPlan"] = "; ".join(
                        g.preparation_guide for g in matching.gaps[:5] if g.preparation_guide
                    )

            except Exception as exc:
                logger.warning("Batch eval failed for %s: %s", bid_ntce_no, exc)
                job["evaluationReason"] = "분석 중 오류가 발생했습니다."

        return job

    jobs = await asyncio.gather(*[_evaluate_one(bid_no) for bid_no in payload.bid_ntce_nos])

    return {
        "jobsCreated": len(jobs),
        "jobs": list(jobs),
    }


# ── 제안서 초안 생성 ──


class ProposalGeneratePayload(BaseModel):
    session_id: str
    bid_notice_id: str


class GeneralChatPayload(BaseModel):
    message: str
    history: list[dict[str, str]] = []


@app.post("/api/chat/general")
@limiter.limit("30/minute")
async def general_chat(payload: GeneralChatPayload, request: Request) -> dict[str, Any]:
    """세션/문서 없이 일반 대화. 공공조달 도우미 역할."""
    message = payload.message.strip()
    if not message:
        raise HTTPException(status_code=400, detail="message가 비어 있습니다.")

    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        return {"answer": "죄송합니다, AI 응답 기능이 현재 비활성화되어 있어요. 좌측 메뉴에서 기능을 선택해주세요."}

    from openai import OpenAI
    client = OpenAI(api_key=api_key, timeout=60)

    system_msg = (
        "당신은 KiraBot, 공공조달·입찰 전문 AI 비서입니다. "
        "사용자의 질문에 친절하고 간결하게 답변하세요. "
        "공고 검색, 문서 분석, 입찰 전략 등에 대해 도움을 줄 수 있습니다. "
        "사용자가 공고를 검색하고 싶어하면 좌측 '공고 검색/분석' 메뉴를 안내하세요. "
        "한국어로 답변하세요. 답변은 3~4문장 이내로 간결하게."
    )

    messages = [{"role": "system", "content": system_msg}]
    _ALLOWED_ROLES = {"user", "assistant"}
    for h in payload.history[-6:]:
        role = h.get("role", "user")
        if role not in _ALLOWED_ROLES:
            role = "user"
        messages.append({"role": role, "content": h.get("content", "")})
    messages.append({"role": "user", "content": message})

    try:
        from openai import APITimeoutError, APIConnectionError, APIStatusError
        import time as _time

        def _do_general_chat_with_retry():
            last_err = None
            for attempt in range(3):
                try:
                    return client.chat.completions.create(
                        model=os.getenv("OPENAI_CHAT_MODEL", "gpt-4o-mini"),
                        messages=messages,
                        max_tokens=300,
                        temperature=0.7,
                    )
                except (APITimeoutError, APIConnectionError) as e:
                    last_err = e
                    if attempt < 2:
                        _time.sleep(1 * (2 ** attempt))
                except APIStatusError as e:
                    if e.status_code in {429, 500, 502, 503} and attempt < 2:
                        last_err = e
                        _time.sleep(1 * (2 ** attempt))
                    else:
                        raise
            raise last_err  # type: ignore[misc]

        resp = await asyncio.to_thread(_do_general_chat_with_retry)
        answer = resp.choices[0].message.content or "응답을 생성하지 못했어요."
    except Exception:
        answer = "AI 응답 생성 중 일시적인 오류가 발생했어요. 잠시 후 다시 시도해주세요."

    return {"answer": answer}


@app.post("/api/proposal/generate")
async def generate_proposal(payload: ProposalGeneratePayload) -> dict[str, Any]:
    """분석된 문서 기반 제안서 초안 생성."""
    session = _get_or_create_session(payload.session_id)

    if session.latest_rfx_analysis is None:
        raise HTTPException(status_code=400, detail="분석된 문서가 없습니다. 먼저 문서를 분석해주세요.")

    # Build analysis text from session's stored RFx analysis
    analysis = session.latest_rfx_analysis
    analysis_text = f"공고명: {analysis.title}\n"
    analysis_text += f"발주기관: {analysis.issuing_org}\n"
    analysis_text += f"문서유형: {analysis.document_type}\n"
    if analysis.budget:
        analysis_text += f"예산: {analysis.budget}\n"
    if analysis.project_period:
        analysis_text += f"사업기간: {analysis.project_period}\n"
    if analysis.requirements:
        analysis_text += "\n자격요건:\n"
        for req in analysis.requirements:
            analysis_text += f"- [{req.category}] {req.description}\n"
    if analysis.evaluation_criteria:
        analysis_text += "\n평가기준:\n"
        for crit in analysis.evaluation_criteria:
            analysis_text += f"- [{crit.category}] {crit.item} ({crit.score}점)\n"

    # Build company summary from RAG engine
    company_text = ""
    try:
        if session.rag_engine.collection.count() > 0:
            company_results = session.rag_engine.search("회사 소개 강점 실적 인증", top_k=5)
            company_text = "\n".join(r.text for r in company_results if r.text)
    except Exception:
        pass

    # Try calling rag_engine if available
    rag_url = os.environ.get("FASTAPI_URL", "http://localhost:8001")
    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(f"{rag_url}/api/generate-proposal", json={
                "bid_text": analysis_text,
                "company_text": company_text,
            })
            if resp.status_code == 200:
                data = resp.json()
                return {"sections": data.get("sections", {})}
    except Exception:
        pass

    # Fallback: generate template sections using OpenAI
    try:
        api_key = _openai_api_key()
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

        system_prompt = """당신은 공공조달 입찰 제안서 작성 전문가입니다.
주어진 공고 분석 결과와 회사 정보를 바탕으로 제안서 초안의 각 섹션을 작성하세요.
반드시 JSON 객체로 응답하세요. 키는 섹션 제목, 값은 섹션 내용입니다.

필수 섹션: "사업 이해 및 목표", "수행 방법론", "추진 일정", "투입 인력 구성", "기대 효과"
각 섹션은 300-500자로 작성하세요. 한국어로 작성하세요."""

        user_prompt = f"""[공고 분석 결과]
{analysis_text}

[회사 정보]
{company_text or '회사 정보 없음'}

위 정보를 바탕으로 제안서 초안을 JSON 형식으로 작성해주세요."""

        response = client.chat.completions.create(
            model=model,
            max_tokens=3000,
            temperature=0.5,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )
        content = (response.choices[0].message.content or "{}").strip()
        sections = json.loads(content)
        if isinstance(sections, dict) and sections:
            return {"sections": sections}
    except Exception as exc:
        logger.warning("OpenAI 제안서 생성 실패: %s", exc)

    # Final fallback: basic template
    title = analysis.title or payload.bid_notice_id
    sections = {
        "사업 이해 및 목표": f"본 제안서는 '{title}' 공고에 대한 제안입니다.\n발주기관의 사업 목적과 요구사항을 분석하여 최적의 수행 방안을 제시합니다.",
        "수행 방법론": "사업 수행을 위한 체계적인 방법론을 적용합니다.\n요구사항 분석, 설계, 구현, 테스트, 안정화의 단계별 접근법을 따릅니다.",
        "추진 일정": "사업 착수 후 체계적인 일정 관리를 통해 기한 내 완료를 목표로 합니다.\n주요 마일스톤을 설정하고 단계별 산출물을 제출합니다.",
        "투입 인력 구성": "프로젝트 관리자(PM)를 중심으로 분야별 전문 인력을 투입합니다.\n각 인력의 역할과 책임을 명확히 정의합니다.",
        "기대 효과": "본 사업의 성공적 수행을 통해 발주기관의 업무 효율성 향상과 서비스 품질 개선이 기대됩니다.",
    }
    return {"sections": sections}


# ── A-lite 제안서 생성 (Layer 1 knowledge-augmented DOCX) ──


class ProposalGenerateV2Payload(BaseModel):
    session_id: str
    total_pages: int = Field(default=50, ge=10, le=200)
    output_format: str = Field(default="docx", pattern="^(docx|hwpx)$")
    company_id: str = "_default"


def _build_rfx_dict(analysis: Any) -> dict[str, Any]:
    """세션의 rfx_analysis에서 rag_engine용 rfx_dict 생성 (공통 헬퍼).

    evaluation_criteria를 item 기반으로 전달:
    - category: 실제 평가 항목명 (ec.item) — 제안서 섹션명이 됨
    - parent_category: 상위 채점 버킷 (ec.category, 예: 기술평가) — 참조용
    - max_score: 배점
    - description: 상세 설명
    """
    criteria = []
    for ec in (analysis.evaluation_criteria or []):
        item_name = (ec.item or "").strip()
        cat_name = (ec.category or "").strip()
        # item이 있으면 실제 평가항목, 없으면 category 그대로 사용
        section_name = item_name if item_name else cat_name
        criteria.append({
            "category": section_name,
            "parent_category": cat_name,
            "max_score": ec.score,
            "description": ec.detail if hasattr(ec, 'detail') else item_name,
        })
    return {
        "title": analysis.title,
        "issuing_org": analysis.issuing_org,
        "budget": analysis.budget,
        "project_period": analysis.project_period,
        "evaluation_criteria": criteria,
        "requirements": [
            {"category": r.category, "description": r.description}
            for r in (analysis.requirements or [])
        ],
        "rfp_text_summary": "",
    }


@app.post("/api/proposal/generate-v2")
async def generate_proposal_v2(payload: ProposalGenerateV2Payload) -> dict[str, Any]:
    """A-lite: Layer 1 knowledge 기반 DOCX 제안서 생성. rag_engine v2 프록시."""
    session = _get_or_create_session(payload.session_id)

    if session.latest_rfx_analysis is None:
        raise HTTPException(status_code=400, detail="분석된 RFP가 없습니다. 먼저 공고를 분석해주세요.")

    rfx_dict = _build_rfx_dict(session.latest_rfx_analysis)

    try:
        return await _proxy_to_rag(
            "POST", "/api/generate-proposal-v2",
            {"rfx_result": rfx_dict, "total_pages": payload.total_pages, "output_format": payload.output_format, "company_id": payload.company_id},
            timeout=300,
        )
    except HTTPException:
        raise  # Re-raise HTTPException as-is
    except Exception as exc:
        logger.error("proposal-v2 proxy error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail="제안서 생성 중 내부 오류가 발생했습니다.") from exc


# ── 제안서 DOCX 다운로드 프록시 ──


_DOWNLOAD_MIME_TYPES: dict[str, str] = {
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".hwpx": "application/x-hwpml",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ".png": "image/png",
}


@app.get("/api/proposal/download/{filename}")
async def download_proposal_proxy(filename: str):
    """Proxy file download from rag_engine (DOCX/HWPX/XLSX/PPTX/PNG)."""
    import re as _re
    if not _re.match(r'^[a-zA-Z0-9가-힣._\-]+\.(docx|hwpx|xlsx|pptx|png)$', filename) or len(filename) > 150:
        raise HTTPException(status_code=400, detail="잘못된 파일명입니다.")

    ext = os.path.splitext(filename)[1].lower()
    media_type = _DOWNLOAD_MIME_TYPES.get(ext, "application/octet-stream")

    fastapi_url = os.environ.get("FASTAPI_URL", "http://localhost:8001")
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(f"{fastapi_url}/api/proposals/download/{filename}")
        if resp.status_code == 200:
            from fastapi.responses import Response
            # Use ASCII-safe fallback for filename, actual UTF-8 filename in filename*
            # RFC 2231: filename (ASCII) + filename* (UTF-8 encoded)
            ascii_filename = "document.bin"  # Safe fallback for old browsers
            return Response(
                content=resp.content,
                media_type=media_type,
                headers={
                    "Content-Disposition": (
                        f'attachment; filename="{ascii_filename}"; '
                        f"filename*=UTF-8''{urllib.parse.quote(filename)}"
                    )
                },
            )
        elif resp.status_code == 404:
            raise HTTPException(status_code=404, detail="파일을 찾을 수 없습니다.")
        else:
            logger.error("Download proxy error: %s %s", resp.status_code, resp.text[:200])
            raise HTTPException(status_code=502, detail="rag_engine 응답 오류")
    except httpx.HTTPError as exc:
        logger.error("rag_engine proxy failed: %s", exc)
        raise HTTPException(status_code=502, detail="rag_engine 연결 실패") from exc


# ── 체크리스트 프록시 ──


class ChecklistProxyPayload(BaseModel):
    session_id: str


@app.post("/api/proposal/checklist")
async def checklist_proxy(payload: ChecklistProxyPayload) -> dict[str, Any]:
    """Proxy checklist extraction from rag_engine using session's rfx_analysis."""
    session = _get_or_create_session(payload.session_id)
    if session.latest_rfx_analysis is None:
        raise HTTPException(status_code=400, detail="분석된 RFP가 없습니다. 먼저 공고를 분석해주세요.")

    rfx_dict = _build_rfx_dict(session.latest_rfx_analysis)

    return await _proxy_to_rag("POST", "/api/checklist", {"rfx_result": rfx_dict}, timeout=60)


# ── Phase 2: WBS / PPT / 실적기술서 프록시 ──


class GenerateWbsProxyPayload(BaseModel):
    session_id: str
    methodology: str = ""
    use_pack: bool = False
    company_id: str = "_default"


@app.post("/api/proposal/generate-wbs")
async def generate_wbs_proxy(payload: GenerateWbsProxyPayload) -> dict[str, Any]:
    """Proxy WBS generation from rag_engine."""
    session = _get_or_create_session(payload.session_id)
    if session.latest_rfx_analysis is None:
        raise HTTPException(status_code=400, detail="분석된 RFP가 없습니다. 먼저 공고를 분석해주세요.")

    rfx_dict = _build_rfx_dict(session.latest_rfx_analysis)

    # P1: 세션 회사 문서 → 생성 브리지
    company_session_context = _extract_session_company_context(
        session, rfx_title=rfx_dict.get("title", ""),
    )

    return await _proxy_to_rag(
        "POST", "/api/generate-wbs",
        {
            "rfx_result": rfx_dict,
            "methodology": payload.methodology,
            "use_pack": payload.use_pack,
            "company_id": payload.company_id,
            "company_session_context": company_session_context,
        },
        timeout=300,
    )


class GeneratePptProxyPayload(BaseModel):
    session_id: str
    duration_min: int = 30
    qna_count: int = 10
    company_id: str = "_default"


@app.post("/api/proposal/generate-ppt")
async def generate_ppt_proxy(payload: GeneratePptProxyPayload) -> dict[str, Any]:
    """Proxy PPT generation from rag_engine."""
    session = _get_or_create_session(payload.session_id)
    if session.latest_rfx_analysis is None:
        raise HTTPException(status_code=400, detail="분석된 RFP가 없습니다. 먼저 공고를 분석해주세요.")

    rfx_dict = _build_rfx_dict(session.latest_rfx_analysis)

    return await _proxy_to_rag(
        "POST", "/api/generate-ppt",
        {"rfx_result": rfx_dict, "duration_min": payload.duration_min, "qna_count": payload.qna_count, "company_id": payload.company_id},
        timeout=300,
    )


class GenerateTrackRecordProxyPayload(BaseModel):
    session_id: str
    company_id: str = "_default"


@app.post("/api/proposal/generate-track-record")
async def generate_track_record_proxy(payload: GenerateTrackRecordProxyPayload) -> dict[str, Any]:
    """Proxy track record document generation from rag_engine."""
    session = _get_or_create_session(payload.session_id)
    if session.latest_rfx_analysis is None:
        raise HTTPException(status_code=400, detail="분석된 RFP가 없습니다. 먼저 공고를 분석해주세요.")

    rfx_dict = _build_rfx_dict(session.latest_rfx_analysis)

    return await _proxy_to_rag("POST", "/api/generate-track-record", {"rfx_result": rfx_dict, "company_id": payload.company_id}, timeout=300)


# ── 회사 DB 온보딩 프록시 ──


async def _proxy_to_rag(method: str, path: str, json_body: dict | None = None, timeout: int = 30) -> dict:
    """Helper to proxy requests to rag_engine."""
    fastapi_url = os.environ.get("FASTAPI_URL", "http://localhost:8001")
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            if method == "GET":
                resp = await client.get(f"{fastapi_url}{path}")
            elif method == "PUT":
                resp = await client.put(f"{fastapi_url}{path}", json=json_body)
            elif method == "DELETE":
                resp = await client.delete(f"{fastapi_url}{path}")
            else:
                resp = await client.post(f"{fastapi_url}{path}", json=json_body)
        if 200 <= resp.status_code < 300:
            return resp.json()
        # Pass through 4xx status codes so frontend gets correct error info
        try:
            detail = resp.json().get("detail", resp.text[:200])
        except Exception:
            detail = resp.text[:200]
        raise HTTPException(status_code=resp.status_code, detail=detail)
    except httpx.HTTPError as exc:
        logger.error("rag_engine proxy failed: %s", exc)
        raise HTTPException(status_code=502, detail="rag_engine 연결 실패") from exc


@app.post("/api/company-db/track-records")
@limiter.limit("10/minute")
async def proxy_add_track_record(payload: dict, request: Request) -> dict[str, Any]:
    """Proxy: add track record to rag_engine company DB."""
    session_id = payload.get("session_id", "")
    if not session_id or session_id not in SESSIONS:
        raise HTTPException(status_code=401, detail="유효한 세션이 필요합니다.")
    return await _proxy_to_rag("POST", "/api/company-db/track-records", payload)


@app.post("/api/company-db/personnel")
@limiter.limit("10/minute")
async def proxy_add_personnel(payload: dict, request: Request) -> dict[str, Any]:
    """Proxy: add personnel to rag_engine company DB."""
    session_id = payload.get("session_id", "")
    if not session_id or session_id not in SESSIONS:
        raise HTTPException(status_code=401, detail="유효한 세션이 필요합니다.")
    return await _proxy_to_rag("POST", "/api/company-db/personnel", payload)


@app.get("/api/company-db/profile")
@limiter.limit("30/minute")
async def proxy_get_company_db_profile(request: Request, company_id: str = "_default") -> dict[str, Any]:
    """Proxy: get company DB profile from rag_engine."""
    return await _proxy_to_rag("GET", f"/api/company-db/profile?company_id={company_id}")


@app.put("/api/company-db/profile")
@limiter.limit("10/minute")
async def proxy_update_company_db_profile(payload: dict, request: Request) -> dict[str, Any]:
    """Proxy: update company DB profile in rag_engine."""
    session_id = payload.get("session_id", "")
    if not session_id or session_id not in SESSIONS:
        raise HTTPException(status_code=401, detail="유효한 세션이 필요합니다.")
    return await _proxy_to_rag("PUT", "/api/company-db/profile", payload)


@app.get("/api/company-db/stats")
@limiter.limit("30/minute")
async def proxy_get_company_db_stats(request: Request, company_id: str = "_default") -> dict[str, Any]:
    """Proxy: get company DB stats from rag_engine."""
    return await _proxy_to_rag("GET", f"/api/company-db/stats?company_id={company_id}")


@app.get("/api/company-db/canonical-id")
@limiter.limit("30/minute")
async def proxy_get_canonical_company_id(request: Request, company_name: str = "") -> dict[str, Any]:
    """Proxy: get canonical company_id from rag_engine."""
    from urllib.parse import quote
    return await _proxy_to_rag("GET", f"/api/company-db/canonical-id?company_name={quote(company_name)}")


@app.get("/api/company-db/track-records")
@limiter.limit("30/minute")
async def proxy_list_track_records(request: Request, company_id: str = "_default") -> dict[str, Any]:
    """Proxy: list track records from rag_engine company DB."""
    return await _proxy_to_rag("GET", f"/api/company-db/track-records?company_id={company_id}")


@app.get("/api/company-db/personnel")
@limiter.limit("30/minute")
async def proxy_list_personnel(request: Request, company_id: str = "_default") -> dict[str, Any]:
    """Proxy: list personnel from rag_engine company DB."""
    return await _proxy_to_rag("GET", f"/api/company-db/personnel?company_id={company_id}")


@app.delete("/api/company-db/items/{doc_id}")
@limiter.limit("10/minute")
async def proxy_delete_company_db_item(doc_id: str, request: Request, company_id: str = "_default", session_id: str = "") -> dict[str, Any]:
    """Proxy: delete company DB item in rag_engine."""
    if not session_id or session_id not in SESSIONS:
        raise HTTPException(status_code=401, detail="유효한 세션이 필요합니다.")
    return await _proxy_to_rag("DELETE", f"/api/company-db/items/{doc_id}?company_id={company_id}")


# ── Profile.md 편집 프록시 ──


@app.get("/api/profile-md")
async def proxy_get_profile_md(company_id: str = "default") -> dict[str, Any]:
    """Proxy: get profile.md sections from rag_engine."""
    cid = urllib.parse.quote(company_id, safe="")
    return await _proxy_to_rag("GET", f"/api/company-profile/md?company_id={cid}")


@app.put("/api/profile-md/section")
async def proxy_update_profile_section(payload: dict) -> dict[str, Any]:
    """Proxy: update a profile.md section in rag_engine."""
    return await _proxy_to_rag("PUT", "/api/company-profile/md/section", payload)


@app.get("/api/profile-md/history")
async def proxy_get_profile_history(company_id: str = "default") -> dict[str, Any]:
    """Proxy: get profile.md version history from rag_engine."""
    cid = urllib.parse.quote(company_id, safe="")
    return await _proxy_to_rag("GET", f"/api/company-profile/md/history?company_id={cid}")


@app.post("/api/profile-md/rollback")
async def proxy_rollback_profile(payload: dict) -> dict[str, Any]:
    """Proxy: rollback profile.md to a specific version in rag_engine."""
    return await _proxy_to_rag("POST", "/api/company-profile/md/rollback", payload)


# ── 제안서 섹션 편집 프록시 ──


def _validate_docx_filename(fn: str) -> str:
    """Defense-in-depth: reject obviously malicious filenames at proxy layer."""
    if ".." in fn or "/" in fn or "\\" in fn or len(fn) > 200:
        raise HTTPException(status_code=400, detail="잘못된 파일명입니다.")
    return fn


@app.get("/api/proposal-sections")
async def proxy_get_proposal_sections(docx_filename: str) -> dict[str, Any]:
    """Proxy: get proposal sections from rag_engine."""
    _validate_docx_filename(docx_filename)
    fn = urllib.parse.quote(docx_filename, safe="")
    return await _proxy_to_rag("GET", f"/api/proposal-sections?docx_filename={fn}")


@app.put("/api/proposal-sections")
async def proxy_update_proposal_section(payload: dict) -> dict[str, Any]:
    """Proxy: update a proposal section in rag_engine."""
    return await _proxy_to_rag("PUT", "/api/proposal-sections", payload)


@app.post("/api/proposal-sections/reassemble")
async def proxy_reassemble_proposal(payload: dict) -> dict[str, Any]:
    """Proxy: reassemble DOCX from edited sections in rag_engine."""
    return await _proxy_to_rag("POST", "/api/proposal-sections/reassemble", payload, timeout=120)


# ── 알림 설정 CRUD ──

ALERT_SETTINGS_DIR = ROOT_DIR / "data" / "alert_settings"


class SaveAlertSettingsPayload(BaseModel):
    session_id: str = Field(min_length=1, max_length=200)
    keywords: list[str] = Field(default_factory=list, max_length=50)
    categories: list[str] = Field(default_factory=list, max_length=20)
    regions: list[str] = Field(default_factory=list, max_length=20)
    minAmt: float | None = None
    maxAmt: float | None = None
    email: str = Field(min_length=1, max_length=320)
    schedule: str = Field(default="daily", max_length=20)


@app.post("/api/alerts/settings")
async def save_alert_settings(payload: SaveAlertSettingsPayload) -> dict[str, Any]:
    """공고 알림 설정 저장."""
    session_id = _sanitize_alert_session_id(payload.session_id)
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id가 필요합니다.")

    settings = {
        "keywords": payload.keywords,
        "categories": payload.categories,
        "regions": payload.regions,
        "minAmt": payload.minAmt,
        "maxAmt": payload.maxAmt,
        "email": payload.email.strip(),
        "schedule": payload.schedule,
    }

    if not settings["email"]:
        raise HTTPException(status_code=400, detail="이메일 주소가 필요합니다.")

    ALERT_SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
    settings_path = ALERT_SETTINGS_DIR / f"{session_id}.json"
    settings_path.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"ok": True}


@app.get("/api/alerts/settings")
async def get_alert_settings(session_id: str = "") -> dict[str, Any]:
    """공고 알림 설정 조회."""
    session_id = _sanitize_alert_session_id(session_id)
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id가 필요합니다.")

    settings_path = ALERT_SETTINGS_DIR / f"{session_id}.json"
    if not settings_path.exists():
        return {"settings": None}

    try:
        data = json.loads(settings_path.read_text(encoding="utf-8"))
        return {"settings": data}
    except Exception:
        return {"settings": None}


@app.delete("/api/alerts/settings")
async def delete_alert_settings(session_id: str = "") -> dict[str, Any]:
    """공고 알림 설정 삭제."""
    session_id = _sanitize_alert_session_id(session_id)
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id가 필요합니다.")

    settings_path = ALERT_SETTINGS_DIR / f"{session_id}.json"
    if settings_path.exists():
        settings_path.unlink()

    return {"ok": True}


@app.delete("/api/alerts/config")
async def delete_alert_config(request: Request, session_id: str = "") -> dict[str, Any]:
    """다중 규칙 기반 알림 설정 + 상태 완전 삭제."""
    _require_username(request)
    session_id = _sanitize_alert_session_id(session_id)

    config_path = ALERT_CONFIG_DIR / f"{session_id}.json"
    if config_path.exists():
        config_path.unlink()

    state_path = ALERT_STATE_DIR / f"{session_id}.json"
    if state_path.exists():
        state_path.unlink()

    # legacy settings도 같이 정리
    legacy_path = ALERT_SETTINGS_DIR / f"{session_id}.json"
    if legacy_path.exists():
        legacy_path.unlink()

    return {"ok": True}


# ── 알림 규칙 기반 설정 (Alert Config) ──

ALERT_CONFIG_DIR = ROOT_DIR / "data" / "alert_configs"
ALERT_STATE_DIR = ROOT_DIR / "data" / "alert_states"


def _get_alert_state(session_id: str) -> dict:
    """알림 발송 상태 조회."""
    ALERT_STATE_DIR.mkdir(parents=True, exist_ok=True)
    state_path = ALERT_STATE_DIR / f"{session_id}.json"
    if state_path.exists():
        try:
            return json.loads(state_path.read_text("utf-8"))
        except Exception:
            pass
    return {}


def _set_alert_state(session_id: str, state: dict):
    """알림 발송 상태 저장 (atomic write)."""
    ALERT_STATE_DIR.mkdir(parents=True, exist_ok=True)
    state_path = ALERT_STATE_DIR / f"{session_id}.json"
    tmp_path = state_path.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), "utf-8")
    os.replace(str(tmp_path), str(state_path))


def _send_email_brevo(to_email: str, subject: str, html: str,
                      attachments: list[tuple[str, bytes]] | None = None) -> bool:
    """Brevo HTTPS API로 이메일 발송."""
    import base64 as _b64
    brevo_key = os.getenv("BREVO_API_KEY", "").strip()
    sender_email = os.getenv("SMTP_EMAIL", "").strip()
    sender_name = os.getenv("SMTP_SENDER_NAME", "키라봇").strip()

    if not brevo_key:
        return False

    payload: dict[str, Any] = {
        "sender": {"email": sender_email or "noreply@kirabot.co.kr", "name": sender_name},
        "to": [{"email": to_email}],
        "subject": subject,
        "htmlContent": html,
    }

    if attachments:
        payload["attachment"] = [
            {"name": fname, "content": _b64.b64encode(data).decode("ascii")}
            for fname, data in attachments
        ]

    try:
        resp = httpx.post(
            "https://api.brevo.com/v3/smtp/email",
            headers={"api-key": brevo_key, "Content-Type": "application/json"},
            json=payload,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            logger.info("Brevo email sent to %s: %s", to_email, subject)
            return True
        else:
            logger.error("Brevo email failed to %s: %s %s", to_email, resp.status_code, resp.text[:200])
            return False
    except Exception as e:
        logger.error("Brevo email error to %s: %s", to_email, e)
        return False


def _send_email_smtp(to_email: str, subject: str, html: str,
                     attachments: list[tuple[str, bytes]] | None = None) -> bool:
    """Gmail SMTP로 HTML 이메일 발송 (로컬/SMTP 가능 환경용)."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders

    smtp_email = os.getenv("SMTP_EMAIL", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "").strip().replace("\xa0", " ")
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com").strip()
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    sender_name = os.getenv("SMTP_SENDER_NAME", "키라봇").strip()

    if not smtp_email or not smtp_password:
        return False

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = f"{sender_name} <{smtp_email}>"
    msg["To"] = to_email

    html_part = MIMEText(html, "html", "utf-8")
    msg.attach(html_part)

    for filename, data in (attachments or []):
        part = MIMEBase("application", "octet-stream")
        part.set_payload(data)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(part)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(smtp_email, smtp_password)
            server.send_message(msg)
        logger.info("SMTP email sent to %s: %s", to_email, subject)
        return True
    except Exception as e:
        logger.error("SMTP email failed to %s: %s", to_email, e)
        return False


def _send_smtp_email(to_email: str, subject: str, html: str,
                     attachments: list[tuple[str, bytes]] | None = None) -> bool:
    """통합 이메일 발송: Brevo API 우선 → SMTP 폴백."""
    # Brevo API 키가 있으면 Brevo 우선 (Railway 등 SMTP 차단 환경)
    if os.getenv("BREVO_API_KEY", "").strip():
        result = _send_email_brevo(to_email, subject, html, attachments)
        if result:
            return True
        logger.warning("Brevo failed, trying SMTP fallback for %s", to_email)

    # SMTP 폴백 (로컬 개발 등)
    return _send_email_smtp(to_email, subject, html, attachments)


def _send_cancellation_email(to_email: str) -> bool:
    """알림 해제 확인 이메일 발송."""
    app_base_url = os.getenv("APP_BASE_URL", "https://kirabot.co.kr").rstrip("/")

    html = f"""<div style="font-family: -apple-system, sans-serif; color: #334155; max-width: 600px;">
  <h2 style="color: #dc2626;">키라봇 공고 알림이 해제되었습니다</h2>
  <p>요청에 따라 공고 알림 발송이 중지되었습니다.</p>
  <p>더 이상 맞춤 공고 알림 메일을 보내지 않습니다.</p>

  <div style="background: #fef2f2; border: 1px solid #fecaca; border-radius: 8px; padding: 16px; margin: 16px 0;">
    <p style="margin: 0; font-size: 14px; color: #991b1b;">
      알림을 다시 받으시려면
      <a href="{app_base_url}/settings/alerts" style="color: #1e40af; text-decoration: underline;">알림 설정 페이지</a>에서
      다시 활성화해주세요.
    </p>
  </div>

  <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 16px 0;" />
  <p style="color: #94a3b8; font-size: 12px;">키라봇 - 공공조달 입찰 자격 분석 플랫폼</p>
</div>"""

    return _send_smtp_email(to_email, "[키라봇] 공고 알림 해제 완료", html)


def _send_confirmation_email(to_email: str, config: dict, is_update: bool = False) -> bool:
    """알림 등록/변경 확인 이메일 발송."""
    app_base_url = os.getenv("APP_BASE_URL", "https://kirabot.co.kr").rstrip("/")

    # 스케줄 설명
    schedule = config.get("schedule", "daily_1")
    hours = config.get("hours", [])
    if schedule == "realtime":
        schedule_desc = "30분마다 확인"
    elif hours:
        hour_strs = [f"{h}시" for h in sorted(hours)]
        count = len(hours)
        schedule_desc = f"매일 {count}회 ({', '.join(hour_strs)})"
    else:
        schedule_desc_map = {
            "daily_1": "매일 1회 (09시)",
            "daily_2": "매일 2회 (09시, 18시)",
            "daily_3": "매일 3회 (09시, 13시, 18시)",
        }
        schedule_desc = schedule_desc_map.get(schedule, "매일 1회 (09시)")

    # 규칙 설명
    rule_items = ""
    for rule in config.get("rules", []):
        if not rule.get("enabled", True):
            continue
        kws = _html_escape(", ".join(rule.get("keywords", [])) or "(전체)")
        amt_parts = []
        if rule.get("minAmt"):
            amt_parts.append(f'{int(rule["minAmt"]):,}원 이상')
        if rule.get("maxAmt"):
            amt_parts.append(f'{int(rule["maxAmt"]):,}원 이하')
        amt_str = _html_escape(" / ".join(amt_parts) if amt_parts else "제한 없음")
        rule_items += f'<li>키워드: <strong>{kws}</strong> / 금액: {amt_str}</li>'

    title = "키라봇 공고 알림 설정이 변경되었습니다" if is_update else "키라봇 공고 알림 등록 완료"
    intro = ("알림 설정이 변경되었습니다. 변경된 설정으로 공고 알림을 보내드립니다." if is_update
             else "알림이 성공적으로 등록되었습니다. 아래 설정에 따라 공고 알림을 보내드립니다.")
    subject = f"[키라봇] 공고 알림 설정 {'변경' if is_update else '등록'} 완료"

    html = f"""<div style="font-family: -apple-system, sans-serif; color: #334155; max-width: 600px;">
  <h2 style="color: #1e40af;">{title}</h2>
  <p>{intro}</p>

  <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px; margin: 16px 0;">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: #475569;">발송 일정</h3>
    <p style="margin: 0; font-size: 15px;"><strong>{schedule_desc}</strong></p>
  </div>

  <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px; margin: 16px 0;">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: #475569;">검색 규칙</h3>
    <ul style="margin: 0; padding-left: 20px;">{rule_items if rule_items else '<li>(규칙 없음)</li>'}</ul>
  </div>

  <div style="background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 8px; padding: 16px; margin: 16px 0;">
    <h3 style="margin: 0 0 8px 0; font-size: 14px; color: #1e40af;">받으실 내용</h3>
    <ul style="margin: 0; padding-left: 20px; color: #334155;">
      <li>매칭 공고 목록 (엑셀 첨부)</li>
      <li>각 공고별 RFP 요약 분석 (사업유형 / 자격조건 / 배점표)</li>
    </ul>
  </div>

  <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 16px 0;" />
  <p style="color: #64748b; font-size: 13px; margin-bottom: 8px;">
    알림 설정을 변경하거나 해제하려면
    <a href="{app_base_url}/settings/alerts" style="color: #1e40af; text-decoration: underline;">여기</a>를
    클릭해주세요.
  </p>
  <p style="color: #94a3b8; font-size: 12px;">키라봇 - 공공조달 입찰 자격 분석 플랫폼</p>
</div>"""

    return _send_smtp_email(to_email, subject, html)


def _sanitize_alert_session_id(raw: str) -> str:
    """Alert config 전용 session_id 검증 (path traversal 방지)."""
    sid = str(raw or "").strip()
    if not sid:
        raise HTTPException(status_code=400, detail="session_id가 필요합니다.")
    if not re.fullmatch(r"[a-zA-Z0-9_\-]{4,128}", sid):
        raise HTTPException(status_code=400, detail="session_id 형식이 올바르지 않습니다.")
    # Resolve 후 prefix 검증으로 path traversal 차단
    ALERT_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    resolved = (ALERT_CONFIG_DIR / f"{sid}.json").resolve()
    if not str(resolved).startswith(str(ALERT_CONFIG_DIR.resolve())):
        raise HTTPException(status_code=400, detail="잘못된 session_id입니다.")
    return sid


@app.post("/api/alerts/config")
async def save_alert_config_endpoint(request: Request, payload: dict) -> dict[str, Any]:
    """다중 규칙 기반 알림 설정 저장. session_id(인증 필요) 또는 email(공개) 지원."""
    # Email-based save (public access for alert registration)
    if "email" in payload and "session_id" not in payload:
        try:
            # Basic validation (email first, then schedule, then rules)
            email = payload.get("email")
            if not email:
                raise ValueError("이메일이 필요합니다.")

            # Email validation
            if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email) or len(email) > 254:
                raise ValueError("유효한 이메일 주소가 필요합니다.")

            # Validate description length
            if "companyProfile" in payload and payload["companyProfile"]:
                # Add type check
                if not isinstance(payload["companyProfile"], dict):
                    raise ValueError("companyProfile 필드가 객체여야 합니다.")

                desc = payload["companyProfile"].get("description", "")
                # Add None check
                if desc is not None and len(desc) > 2000:
                    raise ValueError("회사 설명은 2000자를 초과할 수 없습니다.")

            # Validate rules
            if not isinstance(payload.get("rules"), list):
                raise ValueError("rules 필드가 배열이어야 합니다.")

            for i, rule in enumerate(payload.get("rules", [])):
                keywords = rule.get("keywords", [])

                # Add type check BEFORE empty check
                if not isinstance(keywords, list):
                    raise ValueError(f"규칙 #{i+1}: keywords 필드가 배열이어야 합니다.")

                if not keywords:
                    raise ValueError(f"규칙 #{i+1}: 최소 1개의 키워드가 필요합니다.")
                if len(keywords) > 50:
                    raise ValueError(f"규칙 #{i+1}: 키워드는 최대 50개까지 가능합니다.")
                for j, kw in enumerate(keywords):
                    if not isinstance(kw, str):
                        raise ValueError(f"규칙 #{i+1}: 키워드 #{j+1}은 문자열이어야 합니다.")
                    if len(kw) > 100:
                        raise ValueError(f"규칙 #{i+1}: 개별 키워드는 100자를 초과할 수 없습니다.")

            # Validate schedule
            schedule = payload.get("schedule", "daily_1")
            if schedule not in ["realtime", "daily_1", "daily_2", "daily_3"]:
                raise ValueError("올바른 schedule 값이 필요합니다 (realtime, daily_1, daily_2, daily_3).")

            # Validate hours
            if "hours" in payload and not isinstance(payload["hours"], list):
                raise ValueError("hours 필드가 배열이어야 합니다.")

            # Validate digestTime (HH:MM format)
            digest_time = payload.get("digestTime")
            if digest_time is not None:
                if not isinstance(digest_time, str) or not re.match(r'^([01]\d|2[0-3]):[0-5]\d$', digest_time):
                    raise ValueError("digestTime은 HH:MM 형식이어야 합니다 (00:00~23:59).")

            # Validate digestDays (array of 0-6)
            digest_days = payload.get("digestDays")
            if digest_days is not None:
                if not isinstance(digest_days, list) or len(digest_days) > 7:
                    raise ValueError("digestDays는 최대 7개의 요일(0~6) 배열이어야 합니다.")
                if any(not isinstance(d, int) or d < 0 or d > 6 for d in digest_days):
                    raise ValueError("digestDays 값은 0(일)~6(토) 정수여야 합니다.")

            # Validate maxPerDay (1-200)
            max_per_day = payload.get("maxPerDay")
            if max_per_day is not None:
                if not isinstance(max_per_day, (int, float)) or int(max_per_day) < 1 or int(max_per_day) > 200:
                    raise ValueError("maxPerDay는 1~200 사이 정수여야 합니다.")

            # Validate quietHours
            quiet_hours = payload.get("quietHours")
            if quiet_hours is not None:
                if not isinstance(quiet_hours, dict):
                    raise ValueError("quietHours 필드가 객체여야 합니다.")
                time_re = r'^([01]\d|2[0-3]):[0-5]\d$'
                qh_start = quiet_hours.get("start", "")
                qh_end = quiet_hours.get("end", "")
                if qh_start and not re.match(time_re, qh_start):
                    raise ValueError("quietHours.start는 HH:MM 형식이어야 합니다.")
                if qh_end and not re.match(time_re, qh_end):
                    raise ValueError("quietHours.end는 HH:MM 형식이어야 합니다.")

            try:
                save_alert_config(payload)
                return {"success": True, "message": "알림 설정이 저장되었습니다."}
            except (IOError, OSError) as e:
                logger.error(f"Failed to save alert config: {e}")
                raise HTTPException(status_code=500, detail="설정 저장에 실패했습니다.")
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    # Session-based save (requires auth) - existing logic
    _require_username(request)
    session_id = _sanitize_alert_session_id(payload.get("session_id", ""))

    config = {
        "enabled": payload.get("enabled", True),
        "email": payload.get("email", "").strip(),
        "schedule": payload.get("schedule", "daily_1"),
        "hours": payload.get("hours", []),
        "rules": payload.get("rules", []),
    }

    if not config["email"]:
        raise HTTPException(status_code=400, detail="이메일 주소가 필요합니다.")

    ALERT_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    config_path = ALERT_CONFIG_DIR / f"{session_id}.json"

    is_new = not config_path.exists()
    tmp_path = config_path.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(str(tmp_path), str(config_path))

    # 확인 이메일 (10분 디바운스)
    confirmation_sent = False
    cancellation_sent = False

    if not config.get("enabled", True):
        # 알림 해제 시 취소 확인 이메일 발송
        cancellation_sent = await asyncio.to_thread(_send_cancellation_email, config["email"])
    else:
        state = _get_alert_state(session_id)
        last_conf = state.get("last_confirmation_sent", "")
        now_iso = datetime.now(timezone.utc).isoformat()
        should_send = True
        if last_conf:
            try:
                last_dt = datetime.fromisoformat(last_conf)
                if (datetime.now(timezone.utc) - last_dt).total_seconds() < 600:
                    should_send = False
            except ValueError:
                pass

        if should_send:
            confirmation_sent = await asyncio.to_thread(_send_confirmation_email, config["email"], config, not is_new)
            if confirmation_sent:
                state["last_confirmation_sent"] = now_iso
                _set_alert_state(session_id, state)

    return {"ok": True, "confirmationSent": confirmation_sent, "cancellationSent": cancellation_sent, "isNew": is_new}


@app.post("/api/alerts/preview")
@limiter.limit("10/minute")
async def preview_alert_matches(request: Request, payload: dict) -> dict[str, Any]:
    """규칙 기반 매칭 공고 미리보기 (이메일 발송 없음)."""
    rules = payload.get("rules", [])
    if not isinstance(rules, list) or not rules:
        return {"count": 0, "bids": []}

    all_bids: list[dict] = []
    for rule in rules:
        if not rule.get("enabled", True):
            continue
        regions = rule.get("regions", [])
        raw_categories = rule.get("categories", [])
        cat_codes = [NARA_CATEGORY_CODE.get(c, c) for c in raw_categories] if raw_categories else ["all"]
        region_list = regions if regions else [""]
        for kw in rule.get("keywords", [""]):
            for rgn in region_list:
                for cat in cat_codes:
                    try:
                        _min = rule.get("minAmt")
                        _max = rule.get("maxAmt")
                        results = await nara_search_bids(
                            keywords=kw, category=cat, region=rgn,
                            min_amt=float(_min) if _min else None,
                            max_amt=float(_max) if _max else None,
                            period="1w", exclude_expired=True, page=1, page_size=20,
                        )
                        all_bids.extend(results.get("notices", []))
                    except Exception as e:
                        logger.warning("[preview] search failed kw='%s': %s", kw, e)

    # Deduplicate
    seen: set[str] = set()
    unique: list[dict] = []
    for b in all_bids:
        bid_id = b.get("id", "")
        if bid_id and bid_id not in seen:
            seen.add(bid_id)
            unique.append(b)

    # Return top 5 samples + total count
    samples = unique[:5]
    return {
        "count": len(unique),
        "bids": [
            {
                "id": b.get("id", ""),
                "title": b.get("title", ""),
                "organization": b.get("organization", ""),
                "deadline": b.get("deadline", ""),
                "amount": b.get("amount"),
            }
            for b in samples
        ],
    }


@app.get("/api/alerts/config")
async def get_alert_config_endpoint(request: Request, session_id: str = "", email: str = "") -> dict[str, Any]:
    """다중 규칙 기반 알림 설정 조회. session_id(인증 필요) 또는 email(공개) 지원."""
    # Email-based lookup (public access for alert registration)
    if email:
        try:
            return get_alert_config(email)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    # Session-based lookup (requires auth)
    _require_username(request)
    session_id = _sanitize_alert_session_id(session_id)

    config_path = ALERT_CONFIG_DIR / f"{session_id}.json"
    if not config_path.exists():
        return {"enabled": True, "email": "", "schedule": "daily_1", "hours": [], "rules": []}

    try:
        data = json.loads(config_path.read_text(encoding="utf-8"))
        return data
    except Exception:
        return {"enabled": True, "email": "", "schedule": "daily_1", "hours": [], "rules": []}


# ── Dashboard Summary + Smart Fit Score ──


@app.get("/api/dashboard/summary")
def dashboard_summary(session_id: str = "") -> dict[str, Any]:
    """대시보드 요약 데이터 반환.

    세션에 저장된 분석 결과를 기반으로 totalAnalyzed, goCount 등을 집계한다.
    아직 구현되지 않은 항목(newMatches, deadlineSoon 등)은 0/빈값으로 반환.
    """
    session_id = session_id.strip()
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id가 필요합니다.")

    total_analyzed = 0
    go_count = 0
    recent_searches: list[dict[str, Any]] = []
    smart_fit_top5: list[dict[str, Any]] = []

    normalized = _sanitize_session_id(session_id)
    session = SESSIONS.get(normalized)
    if session is not None:
        # 현재 세션에 분석 결과가 있으면 카운트
        if session.latest_rfx_analysis is not None:
            total_analyzed = 1
        if session.latest_matching_result is not None:
            rec = (session.latest_matching_result.recommendation or "").upper()
            if rec == "GO":
                go_count = 1

    return {
        "newMatches": 0,
        "deadlineSoon": 0,
        "goCount": go_count,
        "totalAnalyzed": total_analyzed,
        "recentSearches": recent_searches,
        "smartFitTop5": smart_fit_top5,
    }


def _compute_keyword_score(keywords: list[str], bid_title: str, max_score: int = 20) -> tuple[int, str]:
    """키워드 중복 기반 점수를 계산한다.

    각 키워드가 bid_title 에 포함되면 가점. 최대 max_score.
    """
    if not keywords or not bid_title:
        return 0, ""

    title_lower = bid_title.lower()
    matched: list[str] = []
    for kw in keywords:
        kw_norm = kw.strip().lower()
        if kw_norm and kw_norm in title_lower:
            matched.append(kw.strip())

    if not matched:
        return 0, "키워드 일치 없음"

    ratio = len(matched) / len(keywords)
    score = min(max_score, round(ratio * max_score))
    detail = f"일치 키워드: {', '.join(matched)} ({len(matched)}/{len(keywords)})"
    return score, detail


@app.post("/api/smart-fit/score")
def smart_fit_score(payload: SmartFitScorePayload) -> dict[str, Any]:
    """Smart Fit 점수 계산.

    - qualification (max 60): 회사문서 + 분석결과 기반 overall_score 반영
    - keywords (max 20): 키워드-공고제목 중복률
    - region (max 10): placeholder (항상 10)
    - experience (max 10): placeholder (항상 5)
    """
    session_id = payload.session_id.strip()
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id가 필요합니다.")

    normalized = _sanitize_session_id(session_id)
    session = SESSIONS.get(normalized)

    # --- qualification (max 60) ---
    qual_score = 0
    qual_detail = "세션 없음"
    if session is not None:
        company_count = int(session.rag_engine.get_stats().get("total_documents", 0))
        if company_count <= 0:
            qual_score = 0
            qual_detail = "회사 문서 미등록"
        elif session.latest_matching_result is not None:
            # overall_score is 0-100, scale to 0-60
            raw = session.latest_matching_result.overall_score
            qual_score = min(60, max(0, round(raw * 60 / 100)))
            qual_detail = (
                f"종합 적합도 {raw:.0f}% "
                f"(충족 {session.latest_matching_result.met_count}, "
                f"부분충족 {session.latest_matching_result.partially_met_count}, "
                f"미충족 {session.latest_matching_result.not_met_count})"
            )
        else:
            qual_score = 0
            qual_detail = "분석 결과 없음 (공고 문서를 분석해주세요)"

    # --- keywords (max 20) ---
    # bid_title: from payload or from session's latest analysis
    bid_title = payload.bid_title.strip()
    if not bid_title and session is not None and session.latest_rfx_analysis is not None:
        bid_title = session.latest_rfx_analysis.title or ""

    kw_score, kw_detail = _compute_keyword_score(payload.keywords, bid_title, max_score=20)

    # --- region (max 10) — placeholder ---
    region_score = 10
    region_detail = "지역 필터 미적용 (기본값)"

    # --- experience (max 10) — placeholder ---
    exp_score = 5
    exp_detail = "경험 평가 미구현 (기본값)"

    total = qual_score + kw_score + region_score + exp_score

    return {
        "totalScore": total,
        "breakdown": {
            "qualification": {"score": qual_score, "maxScore": 60, "details": qual_detail},
            "keywords": {"score": kw_score, "maxScore": 20, "details": kw_detail},
            "region": {"score": region_score, "maxScore": 10, "details": region_detail},
            "experience": {"score": exp_score, "maxScore": 10, "details": exp_detail},
        },
    }


# ── 발주예측 (Forecast) ──


@app.get("/api/forecast/popular")
async def get_popular_agencies() -> dict[str, Any]:
    """인기 발주기관 목록 반환."""
    agencies = [
        "한국도로공사", "한국수자원공사", "한국전력공사", "한국철도공사",
        "한국토지주택공사", "국방부", "행정안전부", "과학기술정보통신부",
        "교육부", "보건복지부", "환경부", "국토교통부",
        "산업통상자원부", "조달청", "방위사업청", "경찰청",
        "소방청", "기상청", "특허청", "관세청",
    ]
    return {"agencies": agencies}


@app.get("/api/forecast/{org_name}")
async def get_org_forecast(org_name: str) -> dict[str, Any]:
    """기관별 입찰 공고 패턴 + 발주계획 데이터 반환."""
    try:
        # 12개월 공고 데이터 조회
        results_12m = await nara_search_bids(
            keywords=org_name,
            category="all",
            period="12m",
            exclude_expired=False,
            page=1,
            page_size=500,
        )

        notices = results_12m.get("notices", [])

        # 월별 집계 (건수 + 금액)
        monthly: dict[str, dict[str, Any]] = {}
        for bid in notices:
            deadline = bid.get("deadlineAt") or ""
            if deadline:
                month_key = deadline[:7]  # "2026-02"
                if month_key not in monthly:
                    monthly[month_key] = {"count": 0, "totalAmt": 0}
                monthly[month_key]["count"] += 1
                # estimatedPrice: "1,234,567원" → 정수
                raw_price = bid.get("estimatedPrice") or ""
                if raw_price:
                    try:
                        amt = int(re.sub(r"[^0-9]", "", raw_price))
                        monthly[month_key]["totalAmt"] += amt
                    except ValueError:
                        pass

        # 카테고리 분포 집계
        category_counts: dict[str, int] = {}
        for bid in notices:
            cat = bid.get("category", "기타")
            category_counts[cat] = category_counts.get(cat, 0) + 1

        # 발주계획(사전공개) 데이터 조회
        order_plans: list[dict[str, Any]] = []
        try:
            plan_result = await nara_search_order_plans(org_name=org_name)
            all_plans = plan_result.get("plans", [])
            # 아직 입찰공고가 안 나온 사업만 필터
            order_plans = [
                p for p in all_plans
                if not p.get("bidNtceNoList") or p.get("ntcePblancYn") != "Y"
            ]
        except Exception as plan_exc:
            logger.warning("발주계획 조회 실패 (org=%s): %s", org_name, plan_exc)

        total = results_12m.get("total", 0)

        return {
            "orgName": org_name,
            "monthlyPattern": monthly,
            "categoryBreakdown": category_counts,
            "recentBids": notices[:20],
            "orderPlans": order_plans,
            "aiInsight": (
                f"{org_name}의 최근 12개월 입찰 공고 패턴입니다. "
                "이 데이터는 참고용이며 실제 발주 계획과 다를 수 있습니다."
            ),
            "total": total,
        }
    except Exception as exc:
        logger.warning("발주예측 조회 실패 (org=%s): %s", org_name, exc)
        return {
            "orgName": org_name,
            "monthlyPattern": {},
            "categoryBreakdown": {},
            "recentBids": [],
            "orderPlans": [],
            "aiInsight": "데이터를 불러오는 중 오류가 발생했습니다.",
            "total": 0,
        }


# ────────────────────────────────────────────────────────────────
# 회사 프로필 CRUD API
# ────────────────────────────────────────────────────────────────

def _require_username(request: Request) -> str:
    """쿠키에서 username 추출. 미인증 시 401."""
    token = str(request.cookies.get(_auth_cookie_name(), "") or "")
    username = str(resolve_user_from_session(token) or "")
    if not username:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다.")
    return username


_COMPANY_PROFILES_DIR = ROOT_DIR / "data" / "company_profiles"

# ── 결제 / 구독 상수 ──
PLAN_PRICES: dict[str, int] = {"free": 0, "pro": 99_000}
_SUBSCRIPTIONS_DIR = ROOT_DIR / "data" / "subscriptions"


def _safe_username_for_path(username: str) -> str:
    """username을 파일시스템 안전한 문자열로 변환."""
    safe = re.sub(r"[^a-zA-Z0-9@._\-]", "_", username)
    safe = safe.replace("..", "_")
    safe = safe.lstrip(".")
    return safe[:100]


def _subscription_path(username: str) -> Path:
    path = _SUBSCRIPTIONS_DIR / f"{_safe_username_for_path(username)}.json"
    resolved = path.resolve()
    if not resolved.is_relative_to(_SUBSCRIPTIONS_DIR.resolve()):
        raise HTTPException(status_code=400, detail="잘못된 사용자명")
    return path


def _load_subscription(username: str) -> dict | None:
    path = _subscription_path(username)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text("utf-8"))
    except Exception:
        return None


def _save_subscription(username: str, sub: dict) -> None:
    _SUBSCRIPTIONS_DIR.mkdir(parents=True, exist_ok=True)
    path = _subscription_path(username)
    sub["updatedAt"] = datetime.now(timezone.utc).isoformat()
    data = json.dumps(sub, ensure_ascii=False, indent=2)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(data, "utf-8")
    os.replace(str(tmp), str(path))  # POSIX atomic rename


def _public_subscription(sub: dict) -> dict:
    """클라이언트 응답용: 민감 정보 제거."""
    return {k: v for k, v in sub.items() if k not in ("billingKey", "username")}


_BILLING_KEY_RE = re.compile(r"^[a-zA-Z0-9_-]{4,256}$")


async def _verify_billing_key_with_portone(billing_key: str) -> bool:
    """PortOne V2 REST API로 빌링키 실재 여부 확인."""
    if not _BILLING_KEY_RE.match(billing_key):
        return False
    secret = os.getenv("PORTONE_API_SECRET", "")
    if not secret:
        raise HTTPException(
            status_code=503, detail="결제 서비스가 구성되지 않았습니다."
        )
    import httpx
    url = f"https://api.portone.io/billing-keys/{billing_key}"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(
                url, headers={"Authorization": f"PortOne {secret}"}
            )
        return resp.status_code == 200
    except Exception:
        logger.exception("PortOne billingKey verification failed")
        return False


def _company_profile_dir(username: str) -> Path:
    path = _COMPANY_PROFILES_DIR / _safe_username_for_path(username)
    if not path.resolve().is_relative_to(_COMPANY_PROFILES_DIR.resolve()):
        raise HTTPException(status_code=400, detail="잘못된 사용자명")
    return path


def _load_company_profile(username: str) -> dict | None:
    path = _company_profile_dir(username) / "profile.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text("utf-8"))
    except Exception:
        logger.warning("profile.json corrupted for user=%s", username)
        return None


def _save_company_profile(username: str, profile: dict) -> None:
    dirp = _company_profile_dir(username)
    dirp.mkdir(parents=True, exist_ok=True)
    profile["updatedAt"] = datetime.now(timezone.utc).isoformat()
    (dirp / "profile.json").write_text(json.dumps(profile, ensure_ascii=False, indent=2), "utf-8")


async def _extract_company_info_llm(text: str) -> dict:
    """OpenAI를 사용하여 문서에서 회사 정보 추출."""
    import openai
    client = openai.OpenAI()

    response = client.chat.completions.create(
        model=os.getenv("OPENAI_CHAT_MODEL", "gpt-4o-mini"),
        messages=[
            {"role": "system", "content": (
                "다음 문서에서 회사 정보를 추출하세요. JSON으로 반환하세요.\n"
                "필드: companyName, businessType, businessNumber, certifications(배열), "
                "regions(배열), employeeCount(숫자), annualRevenue(문자열), "
                "keyExperience(배열), specializations(배열), summary(한줄 요약)"
            )},
            {"role": "user", "content": text[:12000]},
        ],
        response_format={"type": "json_object"},
        temperature=0,
    )
    return json.loads(response.choices[0].message.content or "{}")


async def _analyze_company_writing_style(all_text: str) -> dict | None:
    """Call rag_engine /api/company-db/analyze-style to analyze writing style."""
    # Split into per-document chunks (separated by --- filename --- markers)
    documents = [d.strip() for d in all_text.split("\n--- ") if len(d.strip()) >= 50]
    if not documents:
        documents = [all_text.strip()]
    try:
        data = await _proxy_to_rag("POST", "/api/company-db/analyze-style", {"documents": documents}, timeout=30)
        return data.get("writing_style")
    except Exception as exc:
        logger.warning("Writing style analysis failed: %s", exc)
        return None


def _add_company_docs_to_vectordb(username: str, text: str) -> None:
    """유저별 영구 벡터 컬렉션에 회사 문서 추가."""
    if not text.strip():
        return
    safe = username.replace("/", "_").replace("..", "_")
    collection_name = f"company_{safe}"
    engine = RAGEngine(
        persist_directory=str(ROOT_DIR / "data" / "vectordb"),
        collection_name=collection_name,
    )
    # Split text into chunks and add
    chunks = [text[i:i + 1000] for i in range(0, len(text), 800)]
    for chunk in chunks[:50]:  # Cap at 50 chunks
        engine.add_text_directly(chunk, source_name="company_profile")


_DOC_ID_RE = re.compile(r"^doc_[a-f0-9]{8}$")
_MAX_CHUNKS_PER_DOC = 10
_MAX_TOTAL_INJECT_CHUNKS = 50


def _inject_company_profile_if_needed(session: "WebRuntimeSession", request: Request) -> int:
    """세션에 회사 문서가 없으면, 로그인 유저의 회사 프로필 문서를 자동 주입.

    Returns the company chunk count after injection.
    """
    with session._inject_lock:
        company_chunk_count = int(session.rag_engine.get_stats().get("total_documents", 0))
        if company_chunk_count > 0:
            return company_chunk_count

        try:
            username = _require_username(request)
        except HTTPException:
            return 0

        profile = _load_company_profile(username)
        if not profile or not profile.get("documents"):
            return 0

        docs_dir = _company_profile_dir(username) / "documents"
        if not docs_dir.exists():
            return 0

        parser = DocumentParser()
        total_injected = 0

        for doc in profile["documents"]:
            if total_injected >= _MAX_TOTAL_INJECT_CHUNKS:
                break
            if not isinstance(doc, dict):
                continue
            doc_id = doc.get("id", "")
            if not _DOC_ID_RE.fullmatch(doc_id):
                logger.warning("Skipping invalid doc_id=%r for user=%s", doc_id, username)
                continue

            for f in docs_dir.glob(f"{doc_id}_*"):
                if not f.resolve().is_relative_to(docs_dir.resolve()):
                    continue
                try:
                    parsed = parser.parse(str(f))
                    text = parsed.text[:8000]
                    doc_chunks = [text[i:i + 1000] for i in range(0, len(text), 800)]
                    for chunk in doc_chunks[:_MAX_CHUNKS_PER_DOC]:
                        if total_injected >= _MAX_TOTAL_INJECT_CHUNKS:
                            break
                        session.rag_engine.add_text_directly(chunk, source_name="company_profile")
                        total_injected += 1
                except Exception:
                    logger.warning("Failed to parse company doc %s for user=%s", f.name, username)

        if total_injected > 0:
            logger.info("Auto-injected company profile docs for user=%s (%d chunks)", username, total_injected)

        return company_chunk_count + total_injected


@app.get("/api/company/profile")
def get_company_profile(request: Request) -> dict[str, Any]:
    username = _require_username(request)
    profile = _load_company_profile(username)
    if not profile:
        return {"ok": True, "profile": None}
    return {"ok": True, "profile": profile}


@app.post("/api/company/profile")
async def upload_company_profile_docs(
    request: Request,
    files: list[UploadFile] = File(...),
) -> dict[str, Any]:
    username = _require_username(request)
    dirp = _company_profile_dir(username)
    docs_dir = dirp / "documents"
    docs_dir.mkdir(parents=True, exist_ok=True)

    profile = _load_company_profile(username) or {
        "companyName": "",
        "businessType": "",
        "businessNumber": "",
        "certifications": [],
        "regions": [],
        "employeeCount": None,
        "annualRevenue": "",
        "keyExperience": [],
        "specializations": [],
        "documents": [],
        "aiExtraction": None,
        "lastAnalyzedAt": None,
        "createdAt": datetime.now(timezone.utc).isoformat(),
    }

    MAX_SIZE = 10 * 1024 * 1024  # 10MB (무료)
    allowed_ext = {".pdf", ".doc", ".docx", ".txt", ".md", ".hwp", ".hwpx", ".xlsx", ".xls", ".csv", ".pptx", ".ppt"}
    saved_docs: list[dict] = []
    all_text = ""

    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in allowed_ext:
            raise HTTPException(status_code=400, detail=f"허용되지 않는 파일 형식: {ext}")
        content = await f.read()
        if len(content) > MAX_SIZE:
            raise HTTPException(status_code=400, detail=f"파일 크기 초과 (최대 10MB): {f.filename}")

        doc_id = f"doc_{uuid.uuid4().hex[:8]}"
        safe_name = os.path.basename(f.filename or "document")
        save_path = docs_dir / f"{doc_id}_{safe_name}"
        if not save_path.resolve().is_relative_to(docs_dir.resolve()):
            raise HTTPException(status_code=400, detail=f"잘못된 파일명: {f.filename}")
        save_path.write_bytes(content)

        saved_docs.append({
            "id": doc_id,
            "name": safe_name,
            "uploadedAt": datetime.now(timezone.utc).isoformat(),
            "size": len(content),
        })

        # Extract text using existing DocumentParser
        try:
            parser = DocumentParser()
            parsed = parser.parse(str(save_path))
            text = parsed.text
            all_text += f"\n--- {safe_name} ---\n{text[:8000]}\n"
        except Exception as exc:
            logger.warning("Company doc parse failed (%s): %s", safe_name, exc)

    profile["documents"] = profile.get("documents", []) + saved_docs

    # LLM extraction
    extraction_status = "skipped"  # skipped | success | partial | failed
    filled_fields: list[str] = []

    if all_text.strip():
        try:
            extraction = await _extract_company_info_llm(all_text)
            profile["aiExtraction"] = {
                "summary": extraction.get("summary", ""),
                "extractedAt": datetime.now(timezone.utc).isoformat(),
                "raw": extraction,
            }
            profile["lastAnalyzedAt"] = datetime.now(timezone.utc).isoformat()
            # Auto-fill empty fields
            field_labels = {
                "companyName": "회사명", "businessType": "업종",
                "businessNumber": "사업자번호", "annualRevenue": "연매출",
            }
            for key in ("companyName", "businessType", "businessNumber", "annualRevenue"):
                if not profile.get(key) and extraction.get(key):
                    profile[key] = extraction[key]
                    filled_fields.append(field_labels.get(key, key))
            for key in ("certifications", "regions", "keyExperience", "specializations"):
                existing = set(profile.get(key, []))
                new_items = [v for v in extraction.get(key, []) if v not in existing]
                for v in new_items:
                    existing.add(v)
                if new_items:
                    filled_fields.append(key)
                profile[key] = list(existing)
            if not profile.get("employeeCount") and extraction.get("employeeCount"):
                profile["employeeCount"] = extraction["employeeCount"]
                filled_fields.append("직원 수")
            extraction_status = "success" if filled_fields else "partial"
        except Exception as e:
            logger.warning("Company profile LLM extraction failed: %s", e)
            extraction_status = "failed"
    else:
        extraction_status = "no_text"

    _save_company_profile(username, profile)

    # Also add to permanent vector DB for RAG
    try:
        _add_company_docs_to_vectordb(username, all_text)
    except Exception as e:
        logger.warning("Company vectordb update failed: %s", e)

    # Analyze writing style from uploaded documents via rag_engine
    writing_style = None
    profile_md_generated = False
    if all_text.strip():
        try:
            writing_style = await _analyze_company_writing_style(all_text)
            if writing_style:
                profile["writingStyle"] = writing_style
                _save_company_profile(username, profile)
        except Exception as e:
            logger.warning("Company writing style analysis failed: %s", e)

        # Auto-generate profile.md so PPT/WBS orchestrators can use it
        company_name = profile.get("companyName") or "미설정"
        try:
            # Use server canonical-id endpoint for consistent company_id
            _encoded_name = urllib.parse.quote(company_name, safe="")
            canonical_resp = await _proxy_to_rag("GET", f"/api/company-db/canonical-id?company_name={_encoded_name}", timeout=5)
            canonical_id = canonical_resp.get("company_id", "_default") if isinstance(canonical_resp, dict) else "_default"
        except Exception:
            canonical_id = "_default"
        try:
            documents = [d.strip() for d in all_text.split("\n--- ") if len(d.strip()) >= 50]
            if not documents:
                documents = [all_text.strip()]
            await _proxy_to_rag(
                "POST",
                "/api/company-profile/generate",
                {"company_name": company_name, "documents": documents, "company_id": canonical_id},
                timeout=30,
            )
            profile_md_generated = True
        except Exception as e:
            logger.warning("Auto profile.md generation failed: %s", e)

    return {
        "ok": True,
        "profile": profile,
        "uploadResult": {
            "savedCount": len(saved_docs),
            "extractionStatus": extraction_status,
            "filledFields": filled_fields,
            "writingStyle": writing_style,
            "profileMdGenerated": profile_md_generated,
        },
    }


@app.put("/api/company/profile")
async def update_company_profile(request: Request) -> dict[str, Any]:
    username = _require_username(request)
    body = await request.json()
    profile = _load_company_profile(username)
    if not profile:
        raise HTTPException(status_code=404, detail="회사 프로필이 없습니다.")

    editable = ("companyName", "businessType", "businessNumber", "certifications",
                "regions", "employeeCount", "annualRevenue", "keyExperience", "specializations")
    for key in editable:
        if key in body:
            profile[key] = body[key]

    _save_company_profile(username, profile)
    return {"ok": True, "profile": profile}


@app.delete("/api/company/profile")
def delete_company_profile(request: Request) -> dict[str, Any]:
    username = _require_username(request)
    dirp = _company_profile_dir(username)
    if dirp.exists():
        shutil.rmtree(dirp, ignore_errors=True)
    return {"ok": True}


@app.delete("/api/company/documents/{doc_id}")
def delete_company_document(request: Request, doc_id: str) -> dict[str, Any]:
    username = _require_username(request)
    profile = _load_company_profile(username)
    if not profile:
        raise HTTPException(status_code=404, detail="프로필 없음")

    profile["documents"] = [d for d in profile.get("documents", []) if d["id"] != doc_id]

    # Delete actual file
    docs_dir = _company_profile_dir(username) / "documents"
    for f in docs_dir.glob(f"{doc_id}_*"):
        f.unlink(missing_ok=True)

    _save_company_profile(username, profile)
    return {"ok": True, "profile": profile}


@app.post("/api/company/reanalyze")
@limiter.limit("5/minute")
async def reanalyze_company_profile(request: Request) -> dict[str, Any]:
    username = _require_username(request)
    profile = _load_company_profile(username)
    if not profile or not profile.get("documents"):
        raise HTTPException(status_code=400, detail="등록된 문서가 없습니다.")

    docs_dir = _company_profile_dir(username) / "documents"
    all_text = ""
    for doc in profile["documents"]:
        for f in docs_dir.glob(f"{doc['id']}_*"):
            try:
                parser = DocumentParser()
                parsed = parser.parse(str(f))
                text = parsed.text
                all_text += f"\n--- {doc['name']} ---\n{text[:8000]}\n"
            except Exception as exc:
                logger.warning("Reanalyze parse failed (%s): %s", doc.get("name", ""), exc)

    if not all_text.strip():
        raise HTTPException(status_code=400, detail="문서에서 텍스트를 추출할 수 없습니다.")

    extraction = await _extract_company_info_llm(all_text)
    profile["aiExtraction"] = {
        "summary": extraction.get("summary", ""),
        "extractedAt": datetime.now(timezone.utc).isoformat(),
        "raw": extraction,
    }
    profile["lastAnalyzedAt"] = datetime.now(timezone.utc).isoformat()
    _save_company_profile(username, profile)
    return {"ok": True, "profile": profile}


# ── 알림 이메일 엑셀 발송 ──


async def _analyze_bid_for_alert(bid: dict, api_key: str, bid_temp_dir: str) -> str:
    """단일 공고의 RFP 3-Section 요약 생성. 실패 시 빈 문자열 반환."""
    bid_id = bid.get("id", "")
    if not bid_id:
        return ""
    try:
        # 1. 첨부파일 조회 (detail → e발주 폴백)
        logger.info("  [alert] bid=%s 첨부파일 조회 중...", bid_id)
        attachments = await nara_get_bid_detail_attachments(bid_id, "00")
        if not attachments:
            attachments = await nara_get_bid_attachments(bid_id, "00")
        best = pick_best_attachment(attachments) if attachments else None
        if not best:
            logger.info("  [alert] bid=%s 첨부파일 없음", bid_id)
            return ""

        # 2. 다운로드 (공고별 서브디렉터리로 파일명 충돌 방지)
        sub_dir = os.path.join(bid_temp_dir, bid_id)
        os.makedirs(sub_dir, exist_ok=True)
        logger.info("  [alert] bid=%s 다운로드 중: %s", bid_id, best["fileNm"])
        local_path = await nara_download_attachment(
            best["fileUrl"], sub_dir, fallback_name=best["fileNm"]
        )

        # 3. 파싱 (DocumentParser — LLM 불필요)
        parser = DocumentParser()
        parsed = await asyncio.to_thread(parser.parse, local_path)
        text = parsed.text
        if not text or len(text.strip()) < 100:
            logger.info("  [alert] bid=%s 파싱 텍스트 너무 짧음 (%d자)", bid_id, len(text or ""))
            return ""

        # 4. RFP 3-Section 요약 (LLM)
        logger.info("  [alert] bid=%s RFP 요약 생성 중...", bid_id)
        analyzer = RFxAnalyzer(
            api_key=api_key,
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
        )
        summary = await asyncio.to_thread(analyzer.generate_rfp_summary, text)
        logger.info("  [alert] bid=%s 요약 완료 (%d자)", bid_id, len(summary))
        return summary

    except Exception as exc:
        logger.warning("  [alert] bid=%s 요약 실패: %s", bid_id, exc)
        return ""


def _build_alert_excel(bids: list[dict], summaries: dict[str, str] | None = None) -> bytes:
    """알림 공고 목록 + RFP 요약을 엑셀 바이트로 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io

    wb = Workbook()
    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Sheet 1: 공고 목록 ──
    ws1 = wb.active
    ws1.title = "공고 목록"
    headers = ["구분", "공고명", "수요처", "부서", "예산금액",
               "공고 게시일시", "입찰서 제출일시", "입찰서 마감일시", "낙찰방법", "비고"]
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, bid in enumerate(bids, 2):
        values = [
            bid.get("category", ""),
            bid.get("title", ""),
            bid.get("demandOrg", bid.get("issuingOrg", "")),
            bid.get("department", ""),
            bid.get("estimatedPrice", ""),
            bid.get("publishedAt", ""),
            bid.get("submitStartAt", ""),
            bid.get("deadlineAt", ""),
            bid.get("awardMethod", ""),
            bid.get("url", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val or "")
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for col in ws1.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Sheet 2: RFP 요약 ──
    if summaries:
        ws2 = wb.create_sheet("RFP 요약")
        s_headers = ["공고번호", "공고명", "RFP 요약"]
        s_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for col, h in enumerate(s_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = s_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 2
        for bid in bids:
            bid_id = bid.get("id", "")
            text = summaries.get(bid_id, "") or "(첨부파일 없음 또는 요약 생성 실패)"
            ws2.cell(row=row, column=1, value=bid_id).border = thin_border
            ws2.cell(row=row, column=2, value=bid.get("title", "")).border = thin_border
            c = ws2.cell(row=row, column=3, value=text)
            c.border = thin_border
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 40
        ws2.column_dimensions["C"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


import html as _html_mod


def _html_escape(s: str) -> str:
    """Escape HTML special characters to prevent XSS in emails."""
    return _html_mod.escape(s, quote=True)


def _send_alert_email(to_email: str, subject: str, bids: list[dict],
                      summaries: dict[str, str] | None = None) -> bool:
    """Gmail SMTP로 엑셀 첨부 HTML 이메일 발송."""
    app_base_url = os.getenv("APP_BASE_URL", "https://kirabot.co.kr").rstrip("/")

    excel_bytes = _build_alert_excel(bids, summaries)
    today = datetime.now().strftime("%Y%m%d")

    summary_count = sum(1 for s in (summaries or {}).values() if s) if summaries else 0

    # 공고 목록 + RFP 요약 HTML 생성
    bid_rows = ""
    for i, bid in enumerate(bids, 1):
        bid_id = bid.get("id", "")
        title = _html_escape(bid.get("title", "제목 없음"))
        org = _html_escape(bid.get("organization", ""))
        deadline = _html_escape(bid.get("deadlineAt", "")[:10]) if bid.get("deadlineAt") else ""
        amt = bid.get("estimatedAmount")
        amt_str = _html_escape(f'{int(amt):,}원') if amt else ""
        link = bid.get("link", "")

        title_html = f'<a href="{_html_escape(link)}" style="color:#1e40af;text-decoration:none;">{title}</a>' if link else title

        bid_rows += f'''<div style="border:1px solid #e2e8f0;border-radius:8px;padding:12px;margin:8px 0;">
  <div style="font-size:13px;color:#64748b;">#{i} · {org} {f'· 마감 {deadline}' if deadline else ''} {f'· {amt_str}' if amt_str else ''}</div>
  <div style="font-size:15px;font-weight:600;margin:4px 0;">{title_html}</div>'''

        # RFP 요약 포함
        summary_text = (summaries or {}).get(bid_id, "")
        if summary_text:
            # 마크다운 → 간단한 HTML 변환 (### → bold, - → bullet)
            summary_html = ""
            for line in summary_text.split("\n"):
                line = line.strip()
                if not line:
                    continue
                if line.startswith("###"):
                    summary_html += f'<div style="font-weight:600;font-size:13px;color:#1e40af;margin:6px 0 2px 0;">{_html_escape(line.lstrip("#").strip())}</div>'
                elif line.startswith("- ") or line.startswith("* "):
                    summary_html += f'<div style="font-size:12px;color:#475569;padding-left:12px;">&bull; {_html_escape(line[2:])}</div>'
                elif line.startswith("|"):
                    continue  # 테이블 행은 이메일에서 생략
                else:
                    summary_html += f'<div style="font-size:12px;color:#475569;">{_html_escape(line)}</div>'
            bid_rows += f'<div style="background:#f8fafc;border-radius:4px;padding:8px;margin-top:6px;">{summary_html}</div>'

        bid_rows += '</div>'

    summary_note = f'<p style="color:#059669;font-size:14px;"><strong>{summary_count}건</strong> RFP 요약 포함</p>' if summary_count > 0 else ''

    html_body = f"""<div style="font-family: -apple-system, sans-serif; color: #334155; max-width: 600px;">
  <h2 style="color: #1e40af;">키라봇 공고 알림</h2>
  <p><strong>{len(bids)}건</strong>의 매칭 공고가 발견되었습니다. {summary_note}</p>

  {bid_rows}

  <p style="color: #64748b; font-size: 13px; margin-top: 16px;">상세 데이터는 첨부 엑셀을 확인해주세요.</p>
  <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 16px 0;" />
  <p style="color: #64748b; font-size: 13px;">
    알림 설정을 변경하거나 해제하려면
    <a href="{app_base_url}/settings/alerts" style="color: #1e40af; text-decoration: underline;">여기</a>를
    클릭해주세요.
  </p>
  <p style="color: #94a3b8; font-size: 12px;">이 메일은 키라봇 알림 설정에 의해 자동 발송되었습니다.</p>
</div>"""

    return _send_smtp_email(
        to_email, subject, html_body,
        attachments=[(f"kirabot_alert_{today}.xlsx", excel_bytes)],
    )


async def _execute_alert_send(config: dict, label: str = "alert",
                              session_id: str = "") -> dict[str, Any]:
    """알림 설정 1건에 대해 검색 + 요약 + 이메일 발송. send-now와 스케줄러 공용."""
    import time as _time
    t0 = _time.perf_counter()

    # Step 1: 검색
    all_bids: list[dict] = []
    for rule in config.get("rules", []):
        if not rule.get("enabled", True):
            continue
        regions = rule.get("regions", [])
        raw_categories = rule.get("categories", [])
        # 한글→API코드 변환, 빈 목록이면 ["all"]
        cat_codes = [NARA_CATEGORY_CODE.get(c, c) for c in raw_categories] if raw_categories else ["all"]
        region_list = regions if regions else [""]
        for kw in rule.get("keywords", [""]):
            for rgn in region_list:
                for cat in cat_codes:
                    try:
                        _min = rule.get("minAmt")
                        _max = rule.get("maxAmt")
                        results = await nara_search_bids(
                            keywords=kw, category=cat, region=rgn,
                            min_amt=float(_min) if _min else None,
                            max_amt=float(_max) if _max else None,
                            period="1w", exclude_expired=True, page=1, page_size=50,
                        )
                        all_bids.extend(results.get("notices", []))
                    except Exception as e:
                        logger.warning("[%s] search failed kw='%s' region='%s' cat='%s': %s", label, kw, rgn, cat, e)

    if not all_bids:
        return {"sent": False, "reason": "매칭 공고가 없습니다.", "count": 0, "summaryCount": 0}

    # Step 2: 중복 제거 + 이미 발송한 공고 제외
    seen: set[str] = set()
    unique: list[dict] = []
    for b in all_bids:
        bid_id = b.get("id", "")
        if bid_id and bid_id not in seen:
            seen.add(bid_id)
            unique.append(b)

    # 이전에 발송한 공고 ID 필터링 (중복 메일 방지)
    already_sent: set[str] = set()
    if session_id:
        state = _get_alert_state(session_id)
        already_sent = set(state.get("sent_bid_ids", []))
    new_bids = [b for b in unique if b.get("id", "") not in already_sent]

    if not new_bids:
        return {"sent": False, "reason": "새로운 공고가 없습니다.", "count": 0, "summaryCount": 0}

    # Apply maxPerDay cap
    max_per_day = config.get("maxPerDay", 50)
    if isinstance(max_per_day, (int, float)):
        max_per_day = max(1, min(200, int(max_per_day)))
    else:
        max_per_day = 50

    total_found = len(new_bids)
    new_bids = new_bids[:max_per_day]
    new_bids = new_bids[:10]

    # Step 3: 순차 RFP 요약 (메모리 절약: 1건씩 처리)
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    summaries: dict[str, str] = {}

    if api_key:
        sem = asyncio.Semaphore(1)
        temp_dir = tempfile.mkdtemp(prefix=f"kira_{label}_")

        async def _summarize(bid: dict) -> tuple[str, str]:
            async with sem:
                bid_id = bid.get("id", "")
                summary = await _analyze_bid_for_alert(bid, api_key, temp_dir)
                return (bid_id, summary)

        results = await asyncio.gather(
            *[_summarize(b) for b in new_bids], return_exceptions=True,
        )
        for r in results:
            if isinstance(r, tuple) and r[1]:
                summaries[r[0]] = r[1]

        shutil.rmtree(temp_dir, ignore_errors=True)

    # Step 4: 이메일 발송
    today = datetime.now().strftime("%Y-%m-%d")
    subject = f"[키라봇] {today} 맞춤 공고 알림 ({len(new_bids)}건)"
    sent = await asyncio.to_thread(_send_alert_email, config["email"], subject, new_bids, summaries or None)

    # Step 5: 발송한 공고 ID 기록 (중복 발송 방지)
    if sent and session_id:
        state = _get_alert_state(session_id)
        prev_ids: list[str] = state.get("sent_bid_ids", [])
        new_ids = [b.get("id", "") for b in new_bids if b.get("id")]
        # 기존 목록 뒤에 새 ID 추가 (순서 유지), 중복 제거
        seen = set(prev_ids)
        merged = list(prev_ids)
        for bid_id in new_ids:
            if bid_id not in seen:
                merged.append(bid_id)
                seen.add(bid_id)
        # 최근 500건만 유지 (오래된 것부터 제거)
        state["sent_bid_ids"] = merged[-500:]
        _set_alert_state(session_id, state)

    elapsed = _time.perf_counter() - t0
    logger.info("[%s] %d건, 요약 %d건 (%.1fs)", label, len(new_bids), len(summaries), elapsed)

    return {"sent": sent, "count": len(new_bids), "totalFound": total_found, "summaryCount": len(summaries)}


@app.post("/api/debug/send-now-test")
@limiter.limit("10/minute")
async def debug_send_now_test(request: Request, payload: dict) -> dict[str, Any]:
    """디버그: 알림 즉시 발송 테스트. 관리자 전용."""
    _require_admin(request)
    config = {
        "email": payload.get("email", "").strip(),
        "rules": payload.get("rules", []),
    }
    if not config["email"]:
        return {"ok": False, "error": "email 필요"}
    if not config["rules"]:
        return {"ok": False, "error": "rules 필요"}
    result = await _execute_alert_send(config, label="debug-send-now")
    return {"ok": True, **result}


@app.post("/api/alerts/send-now")
async def send_alert_now(request: Request, payload: dict) -> dict[str, Any]:
    """알림 설정 기반으로 즉시 공고 검색 + RFP 요약 + 이메일 발송."""
    _require_username(request)
    session_id = payload.get("session_id", "").strip()
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id가 필요합니다.")

    config_path = ALERT_CONFIG_DIR / f"{session_id}.json"
    if not config_path.exists():
        raise HTTPException(status_code=400, detail="알림 설정이 없습니다.")
    config = json.loads(config_path.read_text("utf-8"))

    if not config.get("email"):
        raise HTTPException(status_code=400, detail="수신 이메일이 설정되지 않았습니다.")

    result = await _execute_alert_send(config, label=f"send-now:{session_id}", session_id=session_id)
    return {"ok": True, **result}


@app.post("/api/alerts/test-send")
@limiter.limit("3/hour")
async def test_send_alert(request: Request, payload: dict) -> dict[str, Any]:
    """알림 테스트 발송 (이메일 기반, 인증 불필요). 규칙 기반 검색 + 1건만 발송."""
    email = payload.get("email", "").strip()
    if not email or not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
        raise HTTPException(status_code=400, detail="유효한 이메일 주소가 필요합니다.")

    rules = payload.get("rules", [])
    if not rules:
        raise HTTPException(status_code=400, detail="규칙이 필요합니다.")

    config = {"email": email, "rules": rules, "maxPerDay": 1}
    result = await _execute_alert_send(config, label="test-send")
    return {"ok": True, **result}


# ── 알림 스케줄러 ──

_KST = timezone(timedelta(hours=9))


async def _alert_scheduler_loop():
    """정각/:30 정렬 후 30분마다 알림 설정 스캔."""
    await asyncio.sleep(5)  # 서버 부팅 대기

    # 정각 정렬: 다음 :00 또는 :30까지 대기
    now = datetime.now(_KST)
    if now.minute < 30:
        next_check = now.replace(minute=30, second=0, microsecond=0)
    else:
        next_check = (now + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
    wait_seconds = (next_check - now).total_seconds()
    if wait_seconds > 0:
        logger.info("Alert scheduler: waiting %.0fs until %s", wait_seconds, next_check.strftime("%H:%M"))
        await asyncio.sleep(wait_seconds)

    while True:
        try:
            await _check_and_send_scheduled_alerts()
        except Exception as exc:
            logger.error("Alert scheduler error: %s", exc)
        await asyncio.sleep(30 * 60)  # 30분


async def _check_and_send_scheduled_alerts():
    """모든 알림 설정을 순차 스캔 → 발송 시각이면 이메일."""
    if not ALERT_CONFIG_DIR.exists():
        return

    now_kst = datetime.now(_KST)
    current_hour = now_kst.hour

    for config_file in sorted(ALERT_CONFIG_DIR.glob("*.json")):
        try:
            config = json.loads(config_file.read_text("utf-8"))
        except Exception:
            continue

        if not config.get("enabled", True) or not config.get("email"):
            continue

        # 발송 시각 결정
        schedule = config.get("schedule", "daily_1")
        hours = config.get("hours", [])

        # Check quiet hours
        quiet_hours = config.get("quietHours")
        if quiet_hours and quiet_hours.get("enabled"):
            # Check weekend suppression
            if quiet_hours.get("weekendOff") and now_kst.weekday() >= 5:  # 5=Sat, 6=Sun
                continue
            # Check quiet time range
            qh_start = quiet_hours.get("start", "")
            qh_end = quiet_hours.get("end", "")
            if qh_start and qh_end:
                try:
                    start_h, start_m = map(int, qh_start.split(":"))
                    end_h, end_m = map(int, qh_end.split(":"))
                    now_minutes = now_kst.hour * 60 + now_kst.minute
                    start_minutes = start_h * 60 + start_m
                    end_minutes = end_h * 60 + end_m
                    if start_minutes <= end_minutes:
                        if start_minutes <= now_minutes < end_minutes:
                            continue
                    else:  # wraps midnight (e.g., 22:00 ~ 08:00)
                        if now_minutes >= start_minutes or now_minutes < end_minutes:
                            continue
                except (ValueError, TypeError):
                    pass  # invalid format, skip quiet check

        # Use digestTime for determining send hour
        send_hours: list[int] = [9]  # default fallback
        digest_time = config.get("digestTime", "09:00")
        if schedule in ("daily_1", "daily_3") and digest_time:
            try:
                digest_hour = int(digest_time.split(":")[0])
                # For daily_1: send at digest hour only
                if schedule == "daily_1":
                    send_hours = [digest_hour]
                # For daily_3 (weekly): check digestDays too
                elif schedule == "daily_3":
                    digest_days = config.get("digestDays", [1, 2, 3, 4, 5])
                    # Python weekday: Mon=0..Sun=6; our format: Sun=0..Sat=6
                    py_weekday = now_kst.weekday()  # Mon=0..Sun=6
                    our_day = (py_weekday + 1) % 7  # Sun=0..Sat=6
                    if our_day not in digest_days:
                        continue
                    send_hours = [digest_hour]
            except (ValueError, TypeError):
                pass  # fall through to default send_hours=[9]
        elif schedule == "realtime":
            send_hours = list(range(0, 24))
        elif schedule == "hourly":
            send_hours = list(range(8, 20))
        elif hours:
            send_hours = [int(h) for h in hours]
        elif schedule == "daily_2":
            send_hours = [9, 18]

        if current_hour not in send_hours:
            continue

        # 중복 발송 방지 (별도 상태 파일)
        session_id = config_file.stem
        state = _get_alert_state(session_id)
        today_hour_key = now_kst.strftime("%Y-%m-%d-%H")
        if state.get("last_sent") == today_hour_key:
            continue

        # 순차 발송 (설정 간 동시 실행 방지)
        logger.info("스케줄 알림: session=%s, hour=%d", session_id, current_hour)
        try:
            result = await _execute_alert_send(config, label=f"sched:{session_id}", session_id=session_id)
            if not result.get("sent"):
                logger.warning("스케줄 알림 미발송 (session=%s): %s", session_id, result.get("reason", "unknown"))
        except Exception as exc:
            logger.error("스케줄 알림 실패 (session=%s): %s", session_id, exc)
        finally:
            state["last_sent"] = today_hour_key
            _set_alert_state(session_id, state)


# ────────────────────────────────────────────────────────────────
# 결제 / 구독 API
# ────────────────────────────────────────────────────────────────


@app.post("/api/payments/billing-key")
async def register_billing_key(request: Request) -> dict[str, Any]:
    """빌링키 등록 — 프론트에서 PortOne SDK로 발급받은 빌링키를 저장."""
    username = _require_username(request)
    body = await request.json()

    billing_key = str(body.get("billingKey", "")).strip()
    plan = str(body.get("plan", "")).strip().lower()
    card_last4 = str(body.get("cardLast4", "")).strip()

    if not billing_key:
        raise HTTPException(status_code=400, detail="빌링키가 필요합니다.")
    if plan not in PLAN_PRICES or plan == "free":
        raise HTTPException(status_code=400, detail=f"유효하지 않은 플랜: {plan}")

    # PortOne API로 빌링키 유효성 확인
    if not await _verify_billing_key_with_portone(billing_key):
        raise HTTPException(status_code=400, detail="유효하지 않은 빌링키입니다.")

    # 이중 결제 방지: 이미 활성 구독이 있는지 확인
    existing = _load_subscription(username)
    if existing and existing.get("status") == "active" and existing.get("plan") == plan:
        return {
            "ok": True,
            "subscription": _public_subscription(existing),
            "message": "이미 동일 플랜 구독 중입니다.",
        }

    now = datetime.now(timezone.utc)
    sub = {
        "username": username,
        "plan": plan,
        "status": "active",
        "billingKey": billing_key,
        "cardLast4": card_last4,
        "priceKrw": PLAN_PRICES[plan],
        "createdAt": now.isoformat(),
        "currentPeriodStart": now.isoformat(),
        "currentPeriodEnd": (now + timedelta(days=30)).isoformat(),
        "cancelledAt": None,
    }
    _save_subscription(username, sub)
    return {"ok": True, "subscription": _public_subscription(sub)}


@app.get("/api/payments/subscription")
async def get_subscription(request: Request) -> dict[str, Any]:
    """현재 구독 상태 조회."""
    username = _require_username(request)
    sub = _load_subscription(username)
    if not sub:
        return {"ok": True, "subscription": {"plan": "free", "status": "none"}}
    return {"ok": True, "subscription": _public_subscription(sub)}


@app.post("/api/payments/cancel")
async def cancel_subscription(request: Request) -> dict[str, Any]:
    """구독 해지 — 현재 결제 기간 끝까지 유지."""
    username = _require_username(request)
    sub = _load_subscription(username)
    if not sub or sub.get("status") != "active":
        raise HTTPException(status_code=400, detail="활성 구독이 없습니다.")

    sub["status"] = "cancelled"
    sub["cancelledAt"] = datetime.now(timezone.utc).isoformat()
    _save_subscription(username, sub)

    return {"ok": True, "subscription": _public_subscription(sub)}


@app.post("/api/payments/webhook")
async def portone_webhook(request: Request) -> dict[str, Any]:
    """PortOne V2 웹훅 수신 (Svix 서명 검증)."""
    raw_body = await request.body()

    webhook_secret = os.getenv("PORTONE_WEBHOOK_SECRET", "")
    if not webhook_secret:
        logger.warning("PORTONE_WEBHOOK_SECRET 미설정 — 웹훅 거부")
        raise HTTPException(status_code=503, detail="Webhook not configured")

    # Svix 서명 검증: webhook-id, webhook-timestamp, webhook-signature
    wh_id = request.headers.get("webhook-id", "")
    wh_ts = request.headers.get("webhook-timestamp", "")
    wh_sig = request.headers.get("webhook-signature", "")

    if not wh_id or not wh_ts or not wh_sig:
        raise HTTPException(status_code=401, detail="Missing webhook headers")

    # 타임스탬프 ±5분 허용 (리플레이 공격 방지)
    import time as _time
    try:
        ts_int = int(wh_ts)
    except ValueError:
        raise HTTPException(status_code=401, detail="Invalid timestamp")
    if abs(_time.time() - ts_int) > 300:
        raise HTTPException(status_code=401, detail="Webhook timestamp out of range")

    # webhook_secret이 "whsec_" 접두사로 시작하면 제거 후 base64 디코딩
    secret_raw = webhook_secret
    if secret_raw.startswith("whsec_"):
        secret_raw = secret_raw[6:]
    try:
        secret_bytes = base64.b64decode(secret_raw)
    except Exception:
        secret_bytes = secret_raw.encode()

    signed_payload = f"{wh_id}.{wh_ts}.{raw_body.decode()}".encode()
    expected_sig = base64.b64encode(
        hmac_mod.new(secret_bytes, signed_payload, hashlib.sha256).digest()
    ).decode()

    # webhook-signature 헤더는 "v1,<sig>" 형식 (공백 구분 복수 가능)
    valid = any(
        hmac_mod.compare_digest(f"v1,{expected_sig}", part.strip())
        for part in wh_sig.split(" ")
    )
    if not valid:
        logger.warning("PortOne webhook signature mismatch")
        raise HTTPException(status_code=401, detail="Invalid signature")

    try:
        payload = json.loads(raw_body)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    event_type = payload.get("type", "")
    logger.info("PortOne webhook: type=%s", event_type)

    if event_type in ("Transaction.Paid", "BillingKey.Issued"):
        # TODO: Transaction.Paid → currentPeriodEnd 갱신 (정기결제 집행 구현 후)
        pass
    elif event_type == "Transaction.Failed":
        logger.warning("Payment failed: %s", payload)

    return {"ok": True}


@app.post("/api/payments/verify-amount")
async def verify_payment_amount(request: Request) -> dict[str, Any]:
    """결제 금액 서버사이드 검증. 프론트에서 결제 전 호출."""
    username = _require_username(request)
    body = await request.json()
    plan = str(body.get("plan", "")).strip().lower()

    if plan not in PLAN_PRICES:
        raise HTTPException(status_code=400, detail=f"유효하지 않은 플랜: {plan}")

    return {
        "ok": True,
        "plan": plan,
        "amount": PLAN_PRICES[plan],
        "currency": "KRW",
    }


@app.get("/api/pending-knowledge")
async def get_pending_knowledge_proxy(company_id: str, doc_type: str = "proposal") -> dict[str, Any]:
    """Proxy to rag_engine GET /api/pending-knowledge."""
    params = {"company_id": company_id, "doc_type": doc_type}
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.get(f"{RAG_ENGINE_URL}/api/pending-knowledge", params=params)
        r.raise_for_status()
        return r.json()


@app.post("/api/approve-knowledge")
async def approve_knowledge_proxy(request: Request) -> dict[str, Any]:
    """Proxy to rag_engine POST /api/approve-knowledge."""
    body = await request.json()
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.post(f"{RAG_ENGINE_URL}/api/approve-knowledge", json=body)
        r.raise_for_status()
        return r.json()


@app.delete("/api/reject-knowledge")
async def reject_knowledge_proxy(request: Request) -> dict[str, Any]:
    """Proxy to rag_engine DELETE /api/reject-knowledge."""
    body = await request.json()
    async with httpx.AsyncClient(timeout=10.0) as client:
        r = await client.delete(f"{RAG_ENGINE_URL}/api/reject-knowledge", json=body)
        r.raise_for_status()
        return r.json()


@app.get("/{path_name:path}")
def frontend_spa(path_name: str) -> FileResponse:
    """
    React SPA fallback.
    - dist 내 정적 파일이 있으면 해당 파일 반환
    - 없으면 index.html 반환
    """
    dist_index = FRONTEND_DIST_DIR / "index.html"
    if not dist_index.exists():
        raise HTTPException(status_code=404, detail="Frontend build output(dist)가 없습니다.")

    requested = (FRONTEND_DIST_DIR / path_name).resolve()
    if str(requested).startswith(str(FRONTEND_DIST_DIR.resolve())) and requested.is_file():
        return FileResponse(requested)
    return FileResponse(dist_index)


if __name__ == "__main__":
    import uvicorn

    host = os.getenv("WEB_API_HOST", "0.0.0.0").strip() or "0.0.0.0"
    try:
        # Railway sets PORT; fallback to WEB_API_PORT for local dev
        port = int(os.getenv("PORT", os.getenv("WEB_API_PORT", "8000")).strip())
    except ValueError:
        port = 8000

    uvicorn.run("main:app", host=host, port=port, reload=False)
